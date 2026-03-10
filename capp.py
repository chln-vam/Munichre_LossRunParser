import streamlit as st
import os
import json
import tempfile
import csv
import datetime
import re
import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

# ==============================
# FEATURE STORE
# ==============================
FEATURE_STORE_PATH = "feature_store/claims_json"
os.makedirs(FEATURE_STORE_PATH, exist_ok=True)


# ==============================
# UNICODE NORMALIZER
# ==============================
_DASH_TABLE = str.maketrans({
    '\u2013': '-',   # en-dash  –
    '\u2014': '-',   # em-dash  —
    '\u2012': '-',   # figure dash
    '\u2015': '-',   # horizontal bar
    '\u2212': '-',   # minus sign
    '\ufe58': '-',   # small em-dash
    '\ufe63': '-',   # small hyphen-minus
    '\uff0d': '-',   # fullwidth hyphen-minus
    '\u2018': "'",   # left single quote  '
    '\u2019': "'",   # right single quote '
    '\u201c': '"',   # left double quote  "
    '\u201d': '"',   # right double quote "
    '\u00a0': ' ',   # non-breaking space
    '\u202f': ' ',   # narrow no-break space
})

def normalize_str(s: str) -> str:
    """Replace fancy unicode punctuation with plain ASCII equivalents."""
    if not s:
        return s
    return s.translate(_DASH_TABLE)

# ==============================
# PAGE CONFIG
# ==============================
st.set_page_config(layout="wide", page_title="TPA Claims Review Portal")
if "focus_field" not in st.session_state:
    st.session_state.focus_field = None

# ==============================
# STYLING
# ==============================
st.markdown("""
<style>
    .stApp { background-color: #0d1117; color: #c9d1d9; }
    .main-title {
        font-size: 26px; font-weight: 600; padding: 10px 0;
        border-bottom: 1px solid #30363d; margin-bottom: 20px; color: white;
        text-shadow: 0 0 10px rgba(88,166,255,0.7);
    }
    .sheet-title-banner {
        background: #161b22;
        border: 1px solid #30363d;
        border-left: 4px solid #58a6ff;
        border-radius: 6px;
        padding: 10px 16px;
        margin-bottom: 14px;
    }
    .sheet-title-label {
        font-size: 10px;
        color: #8b949e;
        text-transform: uppercase;
        font-weight: bold;
        letter-spacing: 1px;
        margin-bottom: 3px;
    }
    .sheet-title-value {
        font-size: 15px;
        color: #e6edf3;
        font-weight: 600;
    }
    .sheet-subtitle-value {
        font-size: 12px;
        color: #8b949e;
        margin-top: 3px;
    }
    .claim-card {
        background: #161b22; border: 1px solid #30363d; border-radius: 8px;
        padding: 15px; margin-bottom: 10px; cursor: pointer;
        box-shadow: 0 0 0 transparent; transition: all .25s ease;
    }
    .claim-card:hover {
        border-color: #58a6ff;
        box-shadow: 0 0 12px rgba(88,166,255,0.6);
        transform: translateY(-2px);
    }
    .selected-card { border-left: 4px solid #58a6ff; background: #1c2128; box-shadow: 0 0 16px rgba(88,166,255,0.8); }
    .status-text     { font-size: 12px; color: #3fb950; margin-top: 5px; }
    .status-progress { font-size: 12px; color: #d29922; margin-top: 5px; }
    .mid-header-title  { font-size: 26px; font-weight: bold; color: white; margin-bottom: 0px; }
    .mid-header-sub    { font-size: 15px; color: #8b949e; margin-top: 5px; margin-bottom: 5px; }
    .mid-header-status { font-size: 13px; color: #3fb950; margin-bottom: 15px; }
    .incurred-label    { font-size: 14px; color: #8b949e; margin-bottom: 0px; }
    .incurred-amount   { font-size: 26px; font-weight: bold; color: #3fb950; margin-top: 0px; margin-bottom: 20px; }
    div[data-baseweb="input"],
    div[data-baseweb="base-input"],
    div[data-baseweb="select"] {
        background-color: #161b22 !important;
        border: 1px solid #30363d !important;
        border-radius: 6px !important;
    }
    /* Selectbox — read-only dropdown, no typing/editing allowed */
    div[data-baseweb="select"] input {
        caret-color: transparent !important;
        pointer-events: none !important;
        user-select: none !important;
        cursor: pointer !important;
    }
    div[data-baseweb="select"]:focus-within {
        border: 1px solid #30363d !important;
        box-shadow: none !important;
        outline: none !important;
    }
    div[data-baseweb="select"] [data-baseweb="input"] {
        border: none !important;
        box-shadow: none !important;
        outline: none !important;
    }
    /* Remove the red focus ring Streamlit adds */
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div:focus-within,
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div:focus {
        border-color: #30363d !important;
        box-shadow: none !important;
        outline: none !important;
    }
    div[data-baseweb="input"] input {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        background-color: transparent !important;
        font-size: 15px !important;
        padding: 8px 12px !important;
    }
    div[data-baseweb="input"]:has(input:disabled),
    div[data-baseweb="base-input"]:has(input:disabled) {
        background-color: transparent !important;
        border: none !important;
    }
    div[data-baseweb="input"] input:disabled {
        color: #e6edf3 !important;
        -webkit-text-fill-color: #e6edf3 !important;
        cursor: default !important;
        padding-left: 0px !important;
    }
    div[data-testid="stButton"] button {
        background-color: transparent !important;
        color: #8b949e !important;
        border: 1px solid #30363d !important;
        border-radius: 6px !important;
        padding: 2px 8px !important;
        transition: 0.2s;
    }
    div[data-testid="stButton"] button:hover {
        border-color: #58a6ff !important;
        color: #58a6ff !important;
        background-color: #1c2128 !important;
    }
    div[data-testid="stButton"] button:disabled { opacity: 0.3 !important; }
    div[role="dialog"] {
        background-color: #0d1117 !important;
        border: 1px solid #30363d !important;
        border-radius: 10px !important;
    }
    div[role="dialog"] * { color: #c9d1d9 !important; }
    div[role="dialog"] button {
        background-color: transparent !important;
        border: 1px solid #30363d !important;
        color: #8b949e !important;
    }
    div[role="dialog"] button:hover {
        border-color: #58a6ff !important;
        color: #58a6ff !important;
        background-color: #1c2128 !important;
    }
    .format-card {
        background: #161b22; border: 1px solid #30363d; border-radius: 8px;
        padding: 12px; margin-bottom: 8px;
    }
    .format-card.selected {
        border-color: #58a6ff;
        box-shadow: 0 0 10px rgba(88,166,255,0.4);
    }
    .merged-badge {
        display:inline-block; background:#1c2128; border:1px solid #58a6ff;
        border-radius:4px; padding:1px 6px; font-size:10px; color:#58a6ff;
        margin-left:6px; vertical-align:middle;
    }
    .totals-badge {
        display:inline-block; background:#1c2128; border:1px solid #3fb950;
        border-radius:4px; padding:1px 6px; font-size:10px; color:#3fb950;
        margin-left:6px; vertical-align:middle;
    }
    /* Hide invisible form submit buttons (used for Enter-key capture) */
    div[data-testid="stForm"] div[data-testid="stFormSubmitButton"] {
        display: none !important;
    }
    div[data-testid="stForm"] {
        border: none !important;
        padding: 0 !important;
    }
    /* TPA Records panel — full viewport height scroll */
    section[data-testid="stVerticalBlock"] div[data-testid="stVerticalBlockBorderWrapper"] > div {
        max-height: none !important;
    }
    /* Make the left nav scroll container stretch to bottom of page */
    div[data-testid="stVerticalBlock"] > div[style*="overflow"] {
        height: calc(100vh - 180px) !important;
        max-height: calc(100vh - 180px) !important;
    }
    /* Equal size export selection buttons */
    .export-sel-btn div[data-testid="stButton"],
    .export-sel-btn div[data-testid="stButton"] > button {
        height: 38px !important;
        min-height: 38px !important;
        max-height: 38px !important;
    }
    .export-sel-btn div[data-testid="stButton"] > button {
        width: 100% !important;
        padding: 0 !important;
        font-size: 11px !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        line-height: 38px !important;
    }
    .export-sel-btn div[data-testid="stButton"] p,
    .export-sel-btn div[data-testid="stButton"] span {
        margin: 0 !important;
        padding: 0 !important;
        line-height: 38px !important;
        min-height: unset !important;
        height: 38px !important;
        display: inline !important;
    }
</style>
""", unsafe_allow_html=True)


# ==============================
# SHEET NAMES
# ==============================
def get_sheet_names(file_path: str) -> list:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return ["Sheet1"]
    wb    = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    # Put Summary first if present, preserving order of all other sheets
    summary = [n for n in names if n.strip().lower() == "summary"]
    others  = [n for n in names if n.strip().lower() != "summary"]
    return summary + others


# ==============================
# CLASSIFICATION
# ==============================
def classify_sheet(rows):
    text = " ".join(
        str(cell).lower()
        for row in rows[:20]
        for cell in row
        if cell
    )
    if "line of business" in text:
        return "SUMMARY"
    has_claim = any(x in text for x in [
        "claim number", "claim no", "claim #", "claim id",
        "claim ref", "claimant", "file number", "file no"
    ])
    has_loss = any(x in text for x in [
        "loss date", "date of loss", "loss dt", "accident date",
        "occurrence date", "incident date"
    ])
    has_financial = any(x in text for x in [
        "incurred", "paid", "reserve", "outstanding",
        "total paid", "total incurred", "indemnity", "expense"
    ])
    if has_claim and (has_loss or has_financial):
        return "LOSS_RUN"
    if "policy" in text and ("claim" in text or "incurred" in text):
        return "COMMERCIAL_LOSS_RUN"
    if has_claim:
        return "LOSS_RUN"
    return "UNKNOWN"


def extract_merged_cell_metadata(file_path: str, sheet_name: str) -> dict:
    """Extract merged cell ranges and classify them as title/header/data."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return {}

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    merged_info = {}

    for mr in ws.merged_cells.ranges:
        mn_r, mn_c, mx_r, mx_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        cell = ws.cell(mn_r, mn_c)
        val  = str(cell.value).strip() if cell.value else ""

        span_cols = mx_c - mn_c + 1
        span_rows = mx_r - mn_r + 1

        if mn_r <= 3 and span_cols >= 3:
            region_type = "TITLE"
        elif span_cols >= 2 and span_rows == 1:
            region_type = "HEADER"
        else:
            region_type = "DATA"

        key = f"R{mn_r}C{mn_c}"
        merged_info[key] = {
            "value":       val,
            "type":        region_type,
            "row_start":   mn_r,
            "col_start":   mn_c,
            "row_end":     mx_r,
            "col_end":     mx_c,
            "span_cols":   span_cols,
            "span_rows":   span_rows,
            "excel_row":   mn_r,
            "excel_col":   mn_c,
        }

    wb.close()
    return merged_info


def extract_totals_row(file_path: str, sheet_name: str) -> dict:
    """Find and extract totals/summary rows from the sheet."""
    ext    = os.path.splitext(file_path)[1].lower()
    totals = {}

    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        cell_rows = None
    else:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        raw_rows  = [[cell.value for cell in row] for row in ws.iter_rows()]
        cell_rows = [list(row) for row in ws.iter_rows()]
        rows = raw_rows
        wb.close()

    if not rows:
        return totals

    header_row_index = None
    headers = []
    for i, row in enumerate(rows[:20]):
        row_text = " ".join([str(c).lower() for c in row if c])
        if "claim" in row_text and ("date" in row_text or "incurred" in row_text or "paid" in row_text):
            header_row_index = i
            headers = [str(h).strip() if h is not None else f"Column_{j}" for j, h in enumerate(row)]
            break

    if header_row_index is None or not headers:
        return totals

    totals_rows = []
    for r_idx_rel, raw_row in enumerate(rows[header_row_index + 1:]):
        r_idx = header_row_index + 2 + r_idx_rel
        if not any(raw_row):
            continue
        row_text = " ".join([str(c).lower() for c in raw_row if c])
        if any(kw in row_text for kw in ["total", "subtotal", "grand total", "sum", "totals"]):
            row_data = {}
            cell_row = cell_rows[header_row_index + 1 + r_idx_rel] if cell_rows else None
            for c_idx_0, raw_val in enumerate(raw_row):
                if c_idx_0 >= len(headers):
                    continue
                if cell_row and c_idx_0 < len(cell_row):
                    clean_val = format_cell_value_with_fmt(cell_row[c_idx_0])
                    real_col  = cell_row[c_idx_0].column if hasattr(cell_row[c_idx_0], 'column') else c_idx_0 + 1
                else:
                    clean_val = str(raw_val).strip() if raw_val is not None else ""
                    real_col  = c_idx_0 + 1
                if clean_val:
                    row_data[headers[c_idx_0]] = {
                        "value":     clean_val,
                        "excel_row": r_idx,
                        "excel_col": real_col,
                    }
            if row_data:
                totals_rows.append(row_data)

    if totals_rows:
        totals["rows"] = totals_rows
        # Store the excel_row of the first totals row so we can sort it in output
        totals["excel_row"] = totals_rows[0].get(list(totals_rows[0].keys())[0], {}).get("excel_row", 9999)
        agg = {}
        for row_data in totals_rows:
            for field, info in row_data.items():
                try:
                    num = float(str(info["value"]).replace(",", "").replace("$", ""))
                    if field not in agg:
                        agg[field] = 0.0
                    agg[field] += num
                except:
                    pass
        totals["aggregated"] = {k: round(v, 2) for k, v in agg.items()}

    return totals


def format_cell_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S") if value.hour or value.minute else value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return f"{int(value)}.0"
        formatted = f"{value:.10f}".rstrip('0')
        if '.' not in formatted:
            formatted += '.0'
        return formatted
    return normalize_str(str(value).strip())


def _apply_date_number_format(dt, nf: str) -> str:
    """
    Convert a datetime/date to a string by honouring the Excel number_format (nf).
    Uses a single-pass regex replacement (longest tokens first) to avoid
    partial-match bugs like dd→%d leaving a stray 'd'.
    Falls back to MM-DD-YYYY when nf is empty/unrecognised.
    """
    if not nf or nf.lower() in ("general", "@", ""):
        return dt.strftime("%m-%d-%Y")

    # Strip Excel decorator chars that don't affect display
    fmt = re.sub(r'\[.*?\]', '', nf)    # colour/locale brackets e.g. [$-409]
    fmt = re.sub(r'["_*\\]', '', fmt)   # quotes, alignment chars

    result = fmt
    # Protect h-adjacent mm (those mean minutes not months)
    result = re.sub(r'(?i)(?<=h)mm', '__MIN__', result)
    result = re.sub(r'(?i)mm(?=ss)', '__MIN__', result)

    def _tok(m):
        tok = m.group(0).lower()
        return {
            'yyyy': '%Y', 'yy': '%y',
            'mmmm': '%B', 'mmm': '%b', 'mm': '%m', '__min__': '%M', 'm': '%m',
            'dd': '%d', 'd': '%d',
            'hh': '%H', 'h': '%H',
            'ss': '%S', 's': '%S',
            'am/pm': '%p', 'a/p': '%p',
        }.get(tok, m.group(0))

    # Single-pass, longest tokens listed first to avoid partial replacements
    result = re.sub(
        r'(?i)yyyy|yy|mmmm|mmm|__min__|mm|dd|hh|ss|am/pm|a/p|d|h|s|m',
        _tok, result
    )
    try:
        return dt.strftime(result)
    except Exception:
        return dt.strftime("%m-%d-%Y")


def format_cell_value_with_fmt(cell) -> str:
    value = cell.value
    if value is None:
        return ""

    nf = (cell.number_format or "").strip()

    # Dates/datetimes: honour the cell's actual number_format
    if isinstance(value, (datetime.datetime, datetime.date)):
        return _apply_date_number_format(value, nf)

    if isinstance(value, bool):
        return str(value)

    if isinstance(value, (int, float)):
        decimal_places = None

        if nf and nf.lower() not in ("general", "@", ""):
            clean_nf = re.sub(r'[$€£¥"_*\\]', '', nf)
            is_date_fmt = (
                any(x in clean_nf.lower() for x in ['yy', 'mm', 'dd', 'hh', 'ss'])
                and not any(ch in clean_nf for ch in ['0', '#'])
            )
            if not is_date_fmt:
                if '.' in clean_nf:
                    after_dot = clean_nf.split('.')[1]
                    after_dot = re.sub(r'\[.*?\]', '', after_dot)
                    dp = sum(1 for ch in after_dot if ch in '0#')
                    decimal_places = dp
                else:
                    decimal_places = 0

        if decimal_places is not None:
            fval = float(value)
            if decimal_places == 0:
                return str(int(round(fval)))
            return f"{fval:.{decimal_places}f}"

        if isinstance(value, int):
            return str(value)

        fval = float(value)
        remainder = fval - int(fval)
        if remainder == 0.0:
            return f"{fval:.2f}"
        else:
            formatted = f"{fval:.10f}".rstrip('0')
            if '.' not in formatted:
                formatted += '.00'
            elif len(formatted.split('.')[1]) < 2:
                formatted = f"{fval:.2f}"
            return formatted

    return normalize_str(str(value).strip())


def extract_from_excel(file_path, sheet_name):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        if not rows:
            return [], "UNKNOWN"
        sheet_type = classify_sheet(rows)
        return parse_rows(sheet_type, rows)
    else:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        raw_rows   = []
        cell_rows  = []
        for row in ws.iter_rows():
            raw_rows.append([cell.value for cell in row])
            cell_rows.append(list(row))
        wb.close()

        if not raw_rows:
            return [], "UNKNOWN"

        sheet_type = classify_sheet(raw_rows)
        return parse_rows_with_cells(sheet_type, raw_rows, cell_rows)


def parse_rows_with_cells(sheet_type, rows, cell_rows):
    if sheet_type == "SUMMARY":
        header_row_index = None
        for i, row in enumerate(rows[:20]):
            row_text = " ".join([str(c).lower() for c in row if c])
            if "sheet" in row_text and "line of business" in row_text:
                header_row_index = i
                break
        if header_row_index is None:
            return [], sheet_type

        headers = [
            str(h).strip() if h is not None else f"Column_{i}"
            for i, h in enumerate(rows[header_row_index])
        ]
        extracted = []
        for r_idx_rel, (raw_row, cell_row) in enumerate(
            zip(rows[header_row_index + 1:], cell_rows[header_row_index + 1:])
        ):
            r_idx = header_row_index + 2 + r_idx_rel
            if not any(raw_row):
                continue
            row_data = {}
            for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
                if c_idx_0 >= len(headers):
                    continue
                header    = headers[c_idx_0]
                clean_val = format_cell_value_with_fmt(cell)
                real_col  = cell.column if hasattr(cell, 'column') and cell.column else c_idx_0 + 1
                row_data[header] = {
                    "value":    clean_val,
                    "modified": clean_val,
                    "excel_row": r_idx,
                    "excel_col": real_col,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    header_row_index = None
    for i, row in enumerate(rows[:20]):
        row_text = " ".join([str(c).lower() for c in row if c])
        if "claim" in row_text and (
            "date" in row_text or "incurred" in row_text or "paid" in row_text
        ):
            header_row_index = i
            break

    if header_row_index is None:
        return [], sheet_type

    headers = [
        str(h).strip() if h is not None else f"Column_{i}"
        for i, h in enumerate(rows[header_row_index])
    ]
    extracted = []
    data_rows      = rows[header_row_index + 1:]
    data_cell_rows = cell_rows[header_row_index + 1:]

    for r_idx_rel, (raw_row, cell_row) in enumerate(zip(data_rows, data_cell_rows)):
        r_idx = header_row_index + 2 + r_idx_rel
        if not any(raw_row):
            continue
        if any(str(c).lower().strip() in ["totals", "total", "grand total"] for c in raw_row if c):
            break
        row_data = {}
        for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
            if c_idx_0 >= len(headers):
                continue
            header    = headers[c_idx_0]
            clean_val = format_cell_value_with_fmt(cell)
            real_col  = cell.column if hasattr(cell, 'column') and cell.column else c_idx_0 + 1
            row_data[header] = {
                "value":     clean_val,
                "modified":  clean_val,
                "excel_row": r_idx,
                "excel_col": real_col,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


def parse_rows(sheet_type, rows):
    if sheet_type == "SUMMARY":
        header_row_index = None
        for i, row in enumerate(rows[:20]):
            row_text = " ".join([str(c).lower() for c in row if c])
            if "sheet" in row_text and "line of business" in row_text:
                header_row_index = i
                break
        if header_row_index is None:
            return [], sheet_type
        headers = [
            str(h).strip() if h is not None else f"Column_{i}"
            for i, h in enumerate(rows[header_row_index])
        ]
        extracted = []
        for r_idx, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
            if not any(row):
                continue
            row_data = {}
            for c_idx, value in enumerate(row, start=1):
                if c_idx - 1 >= len(headers):
                    continue
                header    = headers[c_idx - 1]
                clean_val = str(value).strip() if value is not None else ""
                row_data[header] = {
                    "value": clean_val, "modified": clean_val,
                    "excel_row": r_idx, "excel_col": c_idx,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    header_row_index = None
    for i, row in enumerate(rows[:20]):
        row_text = " ".join([str(c).lower() for c in row if c])
        if "claim" in row_text and (
            "date" in row_text or "incurred" in row_text or "paid" in row_text
        ):
            header_row_index = i
            break

    if header_row_index is None:
        return [], sheet_type

    headers = [
        str(h).strip() if h is not None else f"Column_{i}"
        for i, h in enumerate(rows[header_row_index])
    ]
    extracted = []
    for r_idx, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
        if not any(row):
            continue
        if any(str(cell).lower().strip() in ["totals", "total", "grand total"] for cell in row if cell):
            break
        row_data = {}
        for c_idx, value in enumerate(row, start=1):
            if c_idx - 1 >= len(headers):
                continue
            header    = headers[c_idx - 1]
            clean_val = str(value).strip() if value is not None else ""
            row_data[header] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": c_idx,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


# ==============================
# EXCEL CELL RENDERER (simple grid)
# ==============================
_THEME_COLORS = {
    0: "FFFFFF", 1: "000000", 2: "EEECE1", 3: "1F497D",
    4: "4F81BD", 5: "C0504D", 6: "9BBB59", 7: "8064A2",
    8: "4BACC6", 9: "F79646",
}


def _resolve_color(color_obj, default="FFFFFF") -> str:
    if color_obj is None:
        return default
    t = color_obj.type
    if t == "rgb":
        rgb = color_obj.rgb or ""
        if len(rgb) == 8 and rgb not in ("00000000", "FF000000"):
            return rgb[2:]
        if len(rgb) == 6:
            return rgb
        return default
    if t == "theme":
        base = _THEME_COLORS.get(color_obj.theme, default)
        tint = color_obj.tint or 0.0
        if tint != 0.0:
            r, g, b = int(base[0:2], 16), int(base[2:4], 16), int(base[4:6], 16)
            if tint > 0:
                r = int(r + (255 - r) * tint)
                g = int(g + (255 - g) * tint)
                b = int(b + (255 - b) * tint)
            else:
                r = int(r * (1 + tint))
                g = int(g * (1 + tint))
                b = int(b * (1 + tint))
            return f"{max(0,min(255,r)):02X}{max(0,min(255,g)):02X}{max(0,min(255,b)):02X}"
        return base
    if t == "indexed":
        indexed_map = {
            0: "000000", 1: "FFFFFF", 2: "FF0000", 3: "00FF00",
            4: "0000FF", 5: "FFFF00", 6: "FF00FF", 7: "00FFFF",
            64: "000000", 65: "FFFFFF",
        }
        return indexed_map.get(color_obj.indexed, default)
    return default


def _col_px(ws, c: int, scale: float = 1.0) -> int:
    letter = get_column_letter(c)
    cd = ws.column_dimensions.get(letter)
    w  = cd.width if (cd and cd.width and cd.width > 0) else 8.43
    return max(20, int(w * 8 * scale))


def _row_px(ws, r: int, scale: float = 1.0) -> int:
    rd = ws.row_dimensions.get(r)
    h  = rd.height if (rd and rd.height and rd.height > 0) else 15.0
    return max(14, int(h * 1.5 * scale))


def render_excel_sheet(excel_path: str, sheet_name: str,
                        scale: float = 1.0) -> tuple:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    col_starts = [0]
    for c in range(1, max_col + 1):
        col_starts.append(col_starts[-1] + _col_px(ws, c, scale))

    row_starts = [0]
    for r in range(1, max_row + 1):
        row_starts.append(row_starts[-1] + _row_px(ws, r, scale))

    img_w = col_starts[-1]
    img_h = row_starts[-1]

    img  = Image.new("RGB", (img_w, img_h), "white")
    draw = ImageDraw.Draw(img, "RGBA")

    merged_master: dict = {}
    for mr in ws.merged_cells.ranges:
        mn_r, mn_c, mx_r, mx_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        for rr in range(mn_r, mx_r + 1):
            for cc in range(mn_c, mx_c + 1):
                merged_master[(rr, cc)] = (mn_r, mn_c, mx_r, mx_c)

    drawn_merges: set = set()

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            merge_info = merged_master.get((r, c))

            if merge_info:
                mn_r, mn_c, mx_r, mx_c = merge_info
                if (mn_r, mn_c) in drawn_merges:
                    continue
                drawn_merges.add((mn_r, mn_c))
                x1 = col_starts[mn_c - 1]; y1 = row_starts[mn_r - 1]
                x2 = col_starts[mx_c];      y2 = row_starts[mx_r]
                cell = ws.cell(mn_r, mn_c)
            else:
                x1 = col_starts[c - 1]; y1 = row_starts[r - 1]
                x2 = col_starts[c];      y2 = row_starts[r]
                cell = ws.cell(r, c)

            bg_hex = "FFFFFF"
            if cell.fill and cell.fill.fill_type == "solid":
                bg_hex = _resolve_color(cell.fill.fgColor, "FFFFFF")

            draw.rectangle([x1, y1, x2 - 1, y2 - 1],
                           fill=f"#{bg_hex}", outline="#CCCCCC", width=1)

            val = cell.value
            if val is not None:
                txt_color = "#000000"
                if cell.font and cell.font.color:
                    fc = _resolve_color(cell.font.color, "000000")
                    if fc.upper() != bg_hex.upper():
                        txt_color = f"#{fc}"

                bold    = bool(cell.font and cell.font.bold)
                text    = format_cell_value_with_fmt(cell) if cell.value is not None else ""
                cell_w  = x2 - x1
                ch_w    = 7 if not bold else 8
                max_chars = max(1, (cell_w - 8) // ch_w)
                if len(text) > max_chars:
                    text = text[:max_chars - 1] + "…"

                draw.text((x1 + 4, y1 + 4), text, fill=txt_color)

    wb.close()
    return img, col_starts, row_starts, merged_master


def get_cell_pixel_bbox(col_starts, row_starts, target_row, target_col,
                        merged_master=None):
    c = max(1, min(target_col, len(col_starts) - 1))
    r = max(1, min(target_row, len(row_starts) - 1))

    if merged_master:
        info = merged_master.get((r, c))
        if info:
            mn_r, mn_c, mx_r, mx_c = info
            x1 = col_starts[mn_c - 1]
            y1 = row_starts[mn_r - 1]
            x2 = col_starts[min(mx_c, len(col_starts) - 1)]
            y2 = row_starts[min(mx_r, len(row_starts) - 1)]
            return x1, y1, x2, y2

    x1 = col_starts[c - 1]
    y1 = row_starts[r - 1]
    x2 = col_starts[min(c, len(col_starts) - 1)]
    y2 = row_starts[min(r, len(row_starts) - 1)]
    return x1, y1, x2, y2


def crop_context(img, x1, y1, x2, y2, pad_x=220, pad_y=160):
    iw, ih = img.size
    cx1 = max(0, x1 - pad_x);  cy1 = max(0, y1 - pad_y)
    cx2 = min(iw, x2 + pad_x); cy2 = min(ih, y2 + pad_y)
    cropped = img.crop((cx1, cy1, cx2, cy2))
    return cropped, x1 - cx1, y1 - cy1, x2 - cx1, y2 - cy1


# ==============================
# EYE POPUP — clean cell view
# ==============================
@st.dialog("Cell View", width="large")
def show_eye_popup(field, info, excel_path, sheet_name):
    st.markdown(f"### 📍 {field}")

    value      = info.get("modified", info["value"])
    target_row = info.get("excel_row")
    target_col = info.get("excel_col")

    col_a, col_b = st.columns([1, 1])
    with col_a:
        st.markdown("**Extracted Value**")
        st.code(value if value else "(empty)")
    with col_b:
        r_lbl = target_row or "?"
        c_lbl = target_col or "?"
        col_letter = get_column_letter(target_col) if target_col else "?"
        st.markdown(f"""
            <div style="padding:10px 0; color:#8b949e; font-size:14px;">
                📌 Cell: <span style="color:#58a6ff; font-weight:bold;">{col_letter}{r_lbl}</span>
                &nbsp;&nbsp;|&nbsp;&nbsp;
                Row <span style="color:#c9d1d9;">{r_lbl}</span> · Col <span style="color:#c9d1d9;">{c_lbl}</span>
            </div>
        """, unsafe_allow_html=True)

    if not target_row or not target_col:
        st.warning("No cell location recorded for this field.")
        return

    ext = os.path.splitext(excel_path)[1].lower()
    if ext == ".csv":
        st.info("Cell preview is not available for CSV files.")
        return

    st.markdown("---")
    st.markdown("**📊 Excel Cell Location**")

    cache_key = f"_rendered_{excel_path}_{sheet_name}"
    with st.spinner("Rendering sheet…"):
        if cache_key not in st.session_state:
            rendered_img, col_starts, row_starts, merged_master = render_excel_sheet(
                excel_path, sheet_name, scale=1.0
            )
            st.session_state[cache_key] = (rendered_img, col_starts, row_starts, merged_master)
        else:
            rendered_img, col_starts, row_starts, merged_master = st.session_state[cache_key]

    try:
        img  = rendered_img.copy()
        draw = ImageDraw.Draw(img, "RGBA")

        x1, y1, x2, y2 = get_cell_pixel_bbox(
            col_starts, row_starts, target_row, target_col, merged_master
        )

        draw.rectangle([x1 + 1, y1 + 1, x2 - 1, y2 - 1], fill=(255, 230, 0, 80))
        draw.rectangle([x1, y1, x2, y2], outline=(255, 180, 0, 255), width=3)
        draw.rectangle([x1 + 3, y1 + 3, x2 - 3, y2 - 3], outline=(255, 255, 255, 160), width=1)

        cropped, _, _, _, _ = crop_context(img, x1, y1, x2, y2, pad_x=300, pad_y=200)

        col_letter = get_column_letter(target_col)
        st.image(
            cropped,
            use_container_width=True,
            caption=f"Cell {col_letter}{target_row}  ·  Value: {value or '(empty)'}"
        )

    except Exception as e:
        st.error(f"Rendering error: {e}")
        import traceback
        st.code(traceback.format_exc())


# ==============================
# FORMAT CONVERTER — Standard JSON only
# (Sequential: titles first by excel_row, then records, then totals)
# ==============================
def to_standard_json(export_data: dict, sheet_meta: dict, totals: dict, merged_meta: dict) -> dict:
    """
    Build a sequential JSON that mirrors the physical order in the Excel sheet:
      1. Title / header merged regions (sorted by excel_row ascending)
      2. Data records (in their original row order)
      3. Totals row(s) at the end
    """
    # --- 1. Titles section: sort merged regions by row then col ---
    titles_section = []
    sorted_merges = sorted(
        [(k, v) for k, v in merged_meta.items() if v.get("value")],
        key=lambda x: (x[1]["row_start"], x[1]["col_start"])
    )
    for key, m in sorted_merges:
        titles_section.append({
            "type":      m["type"],          # TITLE / HEADER / DATA
            "value":     m["value"],
            "excel_row": m["excel_row"],
            "excel_col": m["excel_col"],
            "span_cols": m["span_cols"],
            "span_rows": m["span_rows"],
        })

    # --- 2. Records section ---
    records_section = export_data   # already an ordered dict keyed by claim id

    # --- 3. Totals section ---
    totals_section = {}
    if totals:
        totals_section = {
            "excel_row":  totals.get("excel_row"),
            "rows":       totals.get("rows", []),
            "aggregated": totals.get("aggregated", {}),
        }

    return {
        "exportDate":   datetime.datetime.now().isoformat(),
        "sheetMeta": {
            "sheet_name":    sheet_meta.get("sheet_name"),
            "record_count":  sheet_meta.get("record_count"),
        },
        # Sequential sections mirror Excel top-to-bottom layout
        "titleRows":    titles_section,
        "records":      records_section,
        "totals":       totals_section,
        "recordCount":  len(export_data),
    }


# ==============================
# UTILS
# ==============================
def get_val(claim: dict, keys: list, default: str = "") -> str:
    for pk in keys:
        for k, v in claim.items():
            if pk.lower() in str(k).lower():
                return v["value"] or default
    return default


def detect_claim_id(row, index=None):
    keys = [
        "claim id", "claim_id", "claimid",
        "claim number", "claim no", "claim #",
        "claim ref", "claim reference",
        "file number", "record id"
    ]
    for k, v in row.items():
        name = str(k).lower().replace("_", " ").strip()
        if any(x in name for x in keys):
            val = v.get("modified") or v.get("value")
            if val and str(val).strip():
                return str(val)
    if index is not None:
        return str(index + 1)
    return ""


def clean_duplicate_fields(record: dict) -> dict:
    seen, out = set(), {}
    for k, v in record.items():
        if k.strip() not in seen:
            seen.add(k.strip())
            out[k.strip()] = v
    return out


def save_feature_store(sheet_name: str, data: dict) -> str:
    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(FEATURE_STORE_PATH, f"{sheet_name}_{ts}.json")
    def _sanitize(obj):
        if isinstance(obj, dict): return {k: _sanitize(v) for k, v in obj.items()}
        if isinstance(obj, list): return [_sanitize(i) for i in obj]
        if isinstance(obj, str): return normalize_str(obj)
        return obj
    with open(path, "w") as f:
        json.dump(_sanitize(data), f, indent=2, ensure_ascii=False)
    return path


# ==============================
# MAIN APP
# ==============================
col_title, col_sheet_dropdown = st.columns([4, 1])
with col_title:
    st.markdown('<div class="main-title">🛡️ TPA Claims Review Portal</div>', unsafe_allow_html=True)

uploaded = st.file_uploader("Upload Loss Run Excel/CSV", type=["xlsx", "csv"])

if uploaded:
    if "tmpdir" not in st.session_state:
        st.session_state.tmpdir = tempfile.mkdtemp()

    file_ext   = os.path.splitext(uploaded.name)[1]
    excel_path = os.path.join(st.session_state.tmpdir, f"input{file_ext}")

    if st.session_state.get("last_uploaded") != uploaded.name:
        with open(excel_path, "wb") as f:
            f.write(uploaded.read())
        st.session_state.last_uploaded = uploaded.name
        st.session_state.sheet_names   = get_sheet_names(excel_path)
        st.session_state.sheet_cache   = {}
        st.session_state.selected_idx  = 0
        st.session_state.focus_field   = None
        for key in list(st.session_state.keys()):
            if key.startswith("_rendered_"):
                del st.session_state[key]

    with col_sheet_dropdown:
        st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)
        selected_sheet = st.selectbox(
            "Sheet", st.session_state.sheet_names,
            index=0,
            label_visibility="collapsed"
        )

    st.markdown("<hr style='border-color:#30363d; margin-top:-10px;'>", unsafe_allow_html=True)

    if selected_sheet not in st.session_state.sheet_cache:
        with st.spinner(f"Reading '{selected_sheet}'..."):
            data, sheet_type = extract_from_excel(excel_path, selected_sheet)
            merged_meta      = extract_merged_cell_metadata(excel_path, selected_sheet)
            totals_data      = extract_totals_row(excel_path, selected_sheet)
            st.info(f"Detected Sheet Type: **{sheet_type}** | Merged Regions: **{len(merged_meta)}** | Totals Found: **{'Yes' if totals_data else 'No'}**")
            if not data:
                st.warning(f"No data found in sheet '{selected_sheet}'.")
                st.stop()
            for row in data:
                for fld, inf in row.items():
                    for key in ("value", "modified"):
                        if key in inf and isinstance(inf[key], str):
                            inf[key] = normalize_str(inf[key])
            st.session_state.sheet_cache[selected_sheet] = {
                "data":        data,
                "merged_meta": merged_meta,
                "totals":      totals_data,
            }
            st.session_state.selected_idx = 0
            st.session_state.focus_field  = None

    active      = st.session_state.sheet_cache[selected_sheet]
    data        = active["data"]
    merged_meta = active.get("merged_meta", {})
    totals_data = active.get("totals", {})

    if st.session_state.selected_idx >= len(data):
        st.session_state.selected_idx = 0

    curr_claim = data[st.session_state.selected_idx]

    # ── THREE COLUMN LAYOUT: nav | main | format panel ──
    col_nav, col_main, col_fmt = st.columns([1.2, 3.2, 1.4], gap="large")

    # ── LEFT PANEL ─────────────────────────────────────────────────────
    with col_nav:
        with st.container(height=700, border=False):
            st.markdown("<p style='color:#8b949e; font-weight:bold; font-size:12px; text-transform:uppercase;'>TPA Records</p>", unsafe_allow_html=True)
            for i, row_data in enumerate(data):
                is_sel   = "selected-card" if st.session_state.selected_idx == i else ""
                c_id     = detect_claim_id(row_data, i)
                c_name   = get_val(row_data, ["Insured Name", "Name", "Company", "Claimant", "TPA_NAME"], "Unknown Entity")
                raw_st   = get_val(row_data, ["Status", "CLAIM_STATUS"], "")
                c_status = raw_st or ("Yet to Review" if i == 0 else "In Progress" if i == 1 else "Submitted")
                s_cls    = "status-progress" if "progress" in c_status.lower() or c_status.lower() == "open" else "status-text"
                st.markdown(f"""
                <div class="claim-card {is_sel}">
                    <div style="font-weight:bold;color:white;font-size:15px;">{c_id}</div>
                    <div style="color:#8b949e;font-size:13px;margin-top:2px;">{c_name}</div>
                    <div class="{s_cls}">{c_status}</div>
                </div>""", unsafe_allow_html=True)
                if st.button("Select", key=f"sel_{selected_sheet}_{i}", use_container_width=True):
                    st.session_state.selected_idx = i
                    st.session_state.focus_field  = None
                    st.rerun()

    # ── MIDDLE PANEL ───────────────────────────────────────────────────
    with col_main:

        # ── SHEET TITLE BANNER (from merged cell metadata) ──
        sorted_titles = sorted(
            [(k, v) for k, v in merged_meta.items() if v.get("value")],
            key=lambda x: (x[1]["row_start"], x[1]["col_start"])
        )
        if sorted_titles:
            main_title_val = ""
            sub_title_val  = ""
            for _, m in sorted_titles:
                if m["type"] == "TITLE":
                    if not main_title_val:
                        main_title_val = m["value"]
                    elif not sub_title_val:
                        sub_title_val = m["value"]
            if main_title_val or sub_title_val:
                st.markdown(f"""
                <div class="sheet-title-banner">
                    <div class="sheet-title-label">📄 Sheet Title</div>
                    <div class="sheet-title-value">{main_title_val}</div>
                    {"" if not sub_title_val else f'<div class="sheet-subtitle-value">{sub_title_val}</div>'}
                </div>
                """, unsafe_allow_html=True)

        head_left, head_right = st.columns([3, 1])
        curr_claim_id = detect_claim_id(curr_claim)

        with head_left:
            st.markdown("<p style='color:#8b949e;font-weight:bold;font-size:12px;text-transform:uppercase;'>Review Details</p>", unsafe_allow_html=True)
            h_name   = get_val(curr_claim, ["Insured Name", "Name", "Claimant", "TPA_NAME"], "Unknown Entity")
            h_date   = get_val(curr_claim, ["Loss Date", "Date", "LOSS_DATE"], "N/A")
            h_status = get_val(curr_claim, ["Status", "CLAIM_STATUS"], "Submitted")
            h_total  = get_val(curr_claim, ["Total Incurred", "Incurred", "Total", "Amount", "TOTAL_INCURRED"], "$0")
            st.markdown(f"""
                <div class="mid-header-title">{curr_claim_id}</div>
                <div class="mid-header-sub">{h_name} — {h_date}</div>
                <div class="mid-header-status">{h_status}</div>
                <div class="incurred-label">Total Incurred</div>
                <div class="incurred-amount">{h_total}</div>
            """, unsafe_allow_html=True)

        with head_right:
            st.markdown("<p style='color:#8b949e;font-weight:bold;font-size:12px;text-transform:uppercase;text-align:right;'>Export Selection</p>", unsafe_allow_html=True)
            # Right-aligned equal-size buttons
            st.markdown("""
                <div class="export-sel-btn" style="display:flex;justify-content:flex-end;gap:6px;margin-top:2px;">
            """, unsafe_allow_html=True)
            b1, b2 = st.columns([1, 1])
            with b1:
                if st.button("✔ All", key=f"all_{selected_sheet}_{curr_claim_id}", use_container_width=True):
                    for f in curr_claim:
                        st.session_state[f"chk_{selected_sheet}_{curr_claim_id}_{f}"] = True
                    st.rerun()
            with b2:
                if st.button("✘ None", key=f"none_{selected_sheet}_{curr_claim_id}", use_container_width=True):
                    for f in curr_claim:
                        st.session_state[f"chk_{selected_sheet}_{curr_claim_id}_{f}"] = False
                    st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("<hr style='border-color:#30363d;margin-top:8px;'>", unsafe_allow_html=True)

        # Merged cells info banner
        if merged_meta:
            titles  = [v for v in merged_meta.values() if v["type"] == "TITLE" and v["value"]]
            headers = [v for v in merged_meta.values() if v["type"] == "HEADER" and v["value"]]
            if titles or headers:
                merge_html = "<div style='margin-bottom:10px;'>"
                for t in titles[:3]:
                    merge_html += f'<span class="merged-badge">📌 TITLE: {t["value"][:30]}</span> '
                for h in headers[:4]:
                    merge_html += f'<span class="merged-badge">⊞ HEADER: {h["value"][:20]}</span> '
                merge_html += "</div>"
                st.markdown(merge_html, unsafe_allow_html=True)

        hc = st.columns([2, 2.6, 2.6, 0.6, 0.6, 0.5])
        with hc[0]: st.markdown("**FIELD**")
        with hc[1]: st.markdown("**EXTRACTED VALUE**")
        with hc[2]: st.markdown("**MODIFIED VALUE**")

        for field, info in curr_claim.items():
            ek = f"edit_{selected_sheet}_{curr_claim_id}_{field}"
            xk = f"chk_{selected_sheet}_{curr_claim_id}_{field}"
            mk = f"mod_{selected_sheet}_{curr_claim_id}_{field}"

            if ek not in st.session_state: st.session_state[ek] = False
            if xk not in st.session_state: st.session_state[xk] = True
            if mk not in st.session_state: st.session_state[mk] = info.get("modified", info["value"])

            cl, co, cm, ce, cb, cx = st.columns([2, 2.6, 2.6, 0.9, 0.9, 0.5], gap="small")

            with cl:
                _current_val = st.session_state.get(mk, info.get("modified", info["value"]))
                _is_edited   = _current_val != info["value"]
                _edit_dot    = "<span style='color:#d29922;margin-left:4px;font-size:8px;'>●</span>" if _is_edited else ""
                st.markdown(
                    f"<div style='height:40px;display:flex;align-items:center;"
                    f"color:#c9d1d9;font-size:12px;font-weight:bold;text-transform:uppercase;'>"
                    f"{field}{_edit_dot}</div>", unsafe_allow_html=True)

            with co:
                st.text_input("o", value=info["value"],
                              key=f"orig_{selected_sheet}_{curr_claim_id}_{field}",
                              label_visibility="collapsed", disabled=True)

            with cm:
                if st.session_state[ek]:
                    # Use a form so pressing Enter ALWAYS triggers submission,
                    # even if the value hasn't changed
                    with st.form(key=f"form_{selected_sheet}_{curr_claim_id}_{field}", border=False):
                        nv = st.text_input(
                            "m", value=st.session_state.get(mk, info.get("modified", info["value"])),
                            label_visibility="collapsed"
                        )
                        submitted = st.form_submit_button("", use_container_width=False)
                        if submitted:
                            st.session_state[mk] = nv
                            st.session_state.sheet_cache[selected_sheet]["data"][
                                st.session_state.selected_idx][field]["modified"] = nv
                            st.session_state[ek] = False
                            st.rerun()
                else:
                    nv = st.text_input(
                        "m", key=mk, label_visibility="collapsed",
                        disabled=True
                    )
                # Always keep data store in sync
                st.session_state.sheet_cache[selected_sheet]["data"][
                    st.session_state.selected_idx][field]["modified"] = st.session_state.get(mk, info.get("modified", info["value"]))

            with ce:
                if st.button("👁", key=f"eye_{selected_sheet}_{curr_claim_id}_{field}",
                             use_container_width=True):
                    show_eye_popup(field, info, excel_path, selected_sheet)

            with cb:
                if not st.session_state[ek]:
                    if st.button("✏", key=f"ed_{selected_sheet}_{curr_claim_id}_{field}",
                                 use_container_width=True, help="Edit field"):
                        st.session_state[ek] = True
                        st.rerun()
                else:
                    st.markdown(
                        "<div style='height:38px;display:flex;align-items:center;"
                        "justify-content:center;color:#3fb950;font-size:11px;"
                        "border:1px solid #30363d;border-radius:6px;'>↵</div>",
                        unsafe_allow_html=True
                    )

            with cx:
                st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
                st.checkbox("", key=xk, label_visibility="collapsed")

        # Totals section
        if totals_data:
            st.markdown("<hr style='border-color:#30363d;margin-top:12px;'>", unsafe_allow_html=True)
            st.markdown("**📊 Sheet Totals**")
            agg = totals_data.get("aggregated", {})
            if agg:
                t_cols = st.columns(min(4, len(agg)))
                for idx, (k, v) in enumerate(agg.items()):
                    with t_cols[idx % len(t_cols)]:
                        st.markdown(f"""
                        <div style="background:#161b22;border:1px solid #30363d;border-radius:6px;padding:8px 12px;margin-bottom:6px;">
                            <div style="font-size:11px;color:#8b949e;text-transform:uppercase;">{k}</div>
                            <div style="font-size:16px;font-weight:bold;color:#3fb950;">{v:,.2f}</div>
                        </div>""", unsafe_allow_html=True)

    # ── RIGHT PANEL — FORMAT SELECTOR (Standard JSON only) ─────────────
    with col_fmt:
        st.markdown("<p style='color:#8b949e;font-weight:bold;font-size:12px;text-transform:uppercase;'>Export Format</p>", unsafe_allow_html=True)

        # Only Standard JSON — show as a static selected card
        st.markdown(f"""
            <div style="background:#1c2128;border:1px solid #58a6ff;border-radius:8px;
                        padding:10px 12px;margin-bottom:4px;">
                <div style="font-size:14px;color:white;font-weight:bold;">📄 Standard JSON</div>
                <div style="font-size:11px;color:#8b949e;margin-top:3px;">Raw extracted claims data with titles, records &amp; totals in Excel order</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("<hr style='border-color:#30363d;margin-top:12px;'>", unsafe_allow_html=True)

        # Merged cells panel
        if merged_meta:
            st.markdown("<p style='color:#8b949e;font-weight:bold;font-size:11px;text-transform:uppercase;margin-top:12px;'>Merged Regions</p>", unsafe_allow_html=True)
            sorted_merges = sorted(
                [(k, v) for k, v in merged_meta.items() if v["value"]],
                key=lambda x: (x[1]["row_start"], x[1]["col_start"])
            )
            for key, m in sorted_merges[:8]:
                type_color = "#58a6ff" if m["type"] == "TITLE" else "#d29922" if m["type"] == "HEADER" else "#8b949e"
                st.markdown(f"""
                    <div style="background:#161b22;border:1px solid #30363d;border-radius:6px;
                                padding:6px 10px;margin-bottom:4px;">
                        <div style="font-size:10px;color:{type_color};">{m['type']} · R{m['row_start']}C{m['col_start']}→R{m['row_end']}C{m['col_end']}</div>
                        <div style="font-size:12px;color:#c9d1d9;margin-top:2px;">{m['value'][:35]}</div>
                    </div>""", unsafe_allow_html=True)

        st.markdown("<hr style='border-color:#30363d;margin-top:8px;'>", unsafe_allow_html=True)

        # Export button
        if st.button("⬇ Export as Standard JSON", type="primary", use_container_width=True, key=f"export_{selected_sheet}"):
            export_data = {}
            for i, row in enumerate(data):
                c_id = detect_claim_id(row, i)
                rec  = {}
                for fld, inf in row.items():
                    if st.session_state.get(f"chk_{selected_sheet}_{c_id}_{fld}", True):
                        mk_key   = f"mod_{selected_sheet}_{c_id}_{fld}"
                        live_val = st.session_state.get(mk_key, None)
                        orig     = inf.get("value", "")
                        stored   = inf.get("modified", orig)
                        final_val = live_val if live_val is not None else stored
                        rec[fld] = {
                            "value":        final_val,
                            "original":     orig,
                            "edited":       final_val != orig,
                            "excel_row":    inf.get("excel_row"),
                            "excel_col":    inf.get("excel_col"),
                            "record_index": i,
                        }
                export_data[c_id] = clean_duplicate_fields(rec)

            sheet_meta = {
                "sheet_name":    selected_sheet,
                "record_count":  len(data),
                "merged_regions": merged_meta,
            }

            output = to_standard_json(export_data, sheet_meta, totals_data, merged_meta)
            fname  = f"{selected_sheet}_validated.json"

            def _sanitize_for_json(obj):
                if isinstance(obj, dict):
                    return {k: _sanitize_for_json(v) for k, v in obj.items()}
                if isinstance(obj, list):
                    return [_sanitize_for_json(i) for i in obj]
                if isinstance(obj, str):
                    return normalize_str(obj)
                return obj
            output   = _sanitize_for_json(output)
            json_str = json.dumps(output, indent=2, ensure_ascii=False)
            saved    = save_feature_store(selected_sheet, output)

            st.success("✅ Standard JSON export ready!")
            st.download_button(
                f"📥 Download {fname}",
                data=json_str,
                file_name=fname,
                mime="application/json",
                use_container_width=True,
                key=f"dl_{selected_sheet}_standard"
            )
