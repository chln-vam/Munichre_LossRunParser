import streamlit as st
import os
import json
import tempfile
import subprocess
import shutil
import csv
import openpyxl
import fitz
import io
from PIL import Image, ImageDraw
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.core.credentials import AzureKeyCredential
import datetime


# ==============================
# FEATURE STORE (LOCAL REGISTRY)
# ==============================
FEATURE_STORE_PATH = "feature_store/claims_json"
os.makedirs(FEATURE_STORE_PATH, exist_ok=True)

# ==============================
# PAGE CONFIG
# ==============================
st.set_page_config(layout="wide", page_title="TPA Claims Review Portal")
if "focus_field" not in st.session_state:
    st.session_state.focus_field = None

# ==============================
# CUSTOM STYLING (DARK THEME)
# ==============================
st.markdown("""
<style>
    .stApp { background-color: #0d1117; color: #c9d1d9; }
    .main-title { font-size: 26px; font-weight: 600; padding: 10px 0; border-bottom: 1px solid #30363d; margin-bottom: 20px; color: white; }

    .claim-card {
        background: #161b22; border: 1px solid #30363d; border-radius: 8px;
        padding: 15px; margin-bottom: 10px; transition: 0.3s; cursor: pointer;
        box-shadow: 0 0 0 transparent;
    }
    .claim-card:hover { border-color: #58a6ff; box-shadow: 0 0 12px rgba(88,166,255,0.6); transform: translateY(-2px); }
    .selected-card { border-left: 4px solid #58a6ff; background: #1c2128; box-shadow: 0 0 16px rgba(88,166,255,0.8); }
    .status-text { font-size: 12px; color: #3fb950; margin-top: 5px; }
    .status-progress { font-size: 12px; color: #d29922; margin-top: 5px; }

    .mid-header-title { font-size: 26px; font-weight: bold; color: white; margin-bottom: 0px; }
    .mid-header-sub { font-size: 15px; color: #8b949e; margin-top: 5px; margin-bottom: 5px; }
    .mid-header-status { font-size: 13px; color: #3fb950; margin-bottom: 15px; }
    .incurred-label { font-size: 14px; color: #8b949e; margin-bottom: 0px; }
    .incurred-amount { font-size: 26px; font-weight: bold; color: #3fb950; margin-top: 0px; margin-bottom: 20px; }

    .main-title { text-shadow: 0 0 10px rgba(88,166,255,0.7); }

    div[data-baseweb="input"], div[data-baseweb="base-input"], div[data-baseweb="select"] {
        background-color: #161b22 !important;
        border: 1px solid #30363d !important;
        border-radius: 6px !important;
    }
    div[data-baseweb="input"] input {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        background-color: transparent !important;
        font-size: 15px !important;
        padding: 8px 12px !important;
    }
    div[data-baseweb="input"]:has(input:disabled),
    div[data-baseweb="base-input"]:has(input:disabled) {
        background-color: transparent !important;
        border: none !important;
    }
    div[data-baseweb="input"] input:disabled {
        color: #e6edf3 !important;
        -webkit-text-fill-color: #e6edf3 !important;
        cursor: default !important;
        padding-left: 0px !important;
    }

    div[data-testid="stButton"] button {
        background-color: transparent !important;
        color: #8b949e !important;
        border: 1px solid #30363d !important;
        border-radius: 6px !important;
        padding: 2px 8px !important;
        transition: 0.2s;
    }
    div[data-testid="stButton"] button:hover {
        border-color: #58a6ff !important;
        color: #58a6ff !important;
        background-color: #1c2128 !important;
    }

    div[role="dialog"] {
        background-color: #0d1117 !important;
        border: 1px solid #30363d !important;
        border-radius: 10px !important;
    }
    div[role="dialog"] * { color: #c9d1d9 !important; }
    div[role="dialog"] button {
        background-color: transparent !important;
        border: 1px solid #30363d !important;
        color: #8b949e !important;
    }
    div[role="dialog"] button:hover {
        border-color: #58a6ff !important;
        color: #58a6ff !important;
        background-color: #1c2128 !important;
    }

    .left-scroll-container {
        height: calc(100vh - 140px);
        overflow-y: auto;
        padding-right: 6px;
    }
</style>
""", unsafe_allow_html=True)


# ==============================
# AZURE CONFIG
# ==============================
cfg = st.secrets.get("azureai", {})
ENDPOINT = cfg.get("ENDPOINT")
KEY = cfg.get("KEY")
client = DocumentIntelligenceClient(endpoint=ENDPOINT, credential=AzureKeyCredential(KEY))


# ==============================
# CROSS-PLATFORM HELPERS (replaces win32com)
# ==============================
def get_sheet_names(file_path: str) -> list[str]:
    """Returns sheet names using openpyxl (works on Linux/Windows/Mac)."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return ["Sheet1"]
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def convert_sheet_to_pdf(file_path: str, sheet_name: str, pdf_path: str) -> dict:
    """
    Converts a single Excel sheet (or CSV) to PDF using LibreOffice headless.
    Returns exact_headers dict {col_index: header_name}.
    Requires: libreoffice installed (add to packages.txt on Streamlit Cloud).
    """
    exact_headers = {}
    ext = os.path.splitext(file_path)[1].lower()

    # ── Isolated temp workspace to avoid LibreOffice lock conflicts ──
    work_dir = tempfile.mkdtemp()
    lo_profile = os.path.join(work_dir, "lo_profile")

    try:
        if ext == ".csv":
            # Extract headers from CSV
            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                for row in reader:
                    non_empty = [v for v in row if v.strip()]
                    if len(non_empty) >= 2:
                        for i, v in enumerate(row):
                            if v.strip():
                                exact_headers[i] = v.strip()
                        break

            src = file_path  # LibreOffice can convert CSV directly
            base_name = os.path.splitext(os.path.basename(file_path))[0]

        else:
            # ── Load source workbook, pull headers ──
            wb_src = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb_src[sheet_name]

            for row in ws.iter_rows():
                non_empty = [c for c in row if c.value is not None]
                if len(non_empty) >= 2:
                    for cell in non_empty:
                        exact_headers[cell.column - 1] = str(cell.value).strip()
                    break

            # ── Create isolated workbook for this sheet only ──
            wb_new = openpyxl.Workbook()
            ws_new = wb_new.active
            ws_new.title = sheet_name

            for row in ws.iter_rows(values_only=True):
                ws_new.append([str(v) if v is not None else "" for v in row])

            # Page setup: landscape, fit-to-width (mirrors original win32 settings)
            ws_new.page_setup.orientation = "landscape"
            ws_new.page_setup.fitToWidth = 1
            ws_new.page_setup.fitToHeight = 0
            ws_new.sheet_properties.pageSetUpPr.fitToPage = True
            ws_new.print_options.gridLines = True

            wb_src.close()

            src = os.path.join(work_dir, "sheet.xlsx")
            base_name = "sheet"
            wb_new.save(src)
            wb_new.close()

        # ── LibreOffice conversion ──
        result = subprocess.run(
            [
                "soffice",
                f"--env:UserInstallation=file://{lo_profile}",
                "--headless",
                "--norestore",
                "--convert-to", "pdf",
                "--outdir", work_dir,
                src,
            ],
            capture_output=True,
            text=True,
            timeout=120,
        )

        generated = os.path.join(work_dir, f"{base_name}.pdf")
        if os.path.exists(generated):
            shutil.move(generated, pdf_path)
        else:
            raise RuntimeError(
                f"LibreOffice PDF conversion failed.\nSTDOUT: {result.stdout}\nSTDERR: {result.stderr}"
            )

    finally:
        shutil.rmtree(work_dir, ignore_errors=True)

    return exact_headers


# ==============================
# ADI HELPERS
# ==============================
def get_cell_confidence(cell, result):
    if not cell.spans:
        return 0.98
    start = cell.spans[0].offset
    end = start + cell.spans[0].length
    scores = []
    for page in result.pages:
        for w in page.words:
            if w.span.offset >= start and (w.span.offset + w.span.length) <= end:
                word_content = w.content.strip()
                if len(word_content) > 1 or word_content.isalnum():
                    scores.append(w.confidence)
    if not scores:
        return 0.96
    avg_conf = sum(scores) / len(scores)
    boosted = avg_conf + ((1.0 - avg_conf) * 0.4)
    return round(min(0.99, boosted), 3)


def extract_from_adi(pdf_path, exact_headers):
    with open(pdf_path, "rb") as f:
        poller = client.begin_analyze_document(model_id="prebuilt-layout", body=f.read())
    result = poller.result()

    if not result.tables:
        return None, None

    extracted = []
    global_headers_map = None

    for table in result.tables:
        row_counts = {}
        for cell in table.cells:
            row_counts.setdefault(cell.row_index, 0)
            if cell.content and cell.content.strip():
                row_counts[cell.row_index] += 1

        header_row = None
        for r, count in row_counts.items():
            if count >= 4:
                header_row = r
                break

        if header_row is None:
            continue

        detected_headers = {
            cell.column_index: cell.content
            for cell in table.cells
            if cell.row_index == header_row
        }

        local_headers_map = detected_headers

        if local_headers_map and not global_headers_map:
            global_headers_map = local_headers_map

        if not local_headers_map and global_headers_map:
            active_headers_map = global_headers_map
            start_row = 0
        else:
            active_headers_map = local_headers_map if local_headers_map else global_headers_map
            start_row = header_row + 1

        if not active_headers_map:
            continue

        for r in range(start_row, table.row_count):
            row_data = {}
            for c_idx, header_text in active_headers_map.items():
                if exact_headers and c_idx in exact_headers:
                    header_text = exact_headers[c_idx]
                elif not header_text or str(header_text).strip() == "":
                    header_text = f"Column_{c_idx}"

                cell = next(
                    (cl for cl in table.cells if cl.row_index == r and cl.column_index == c_idx),
                    None,
                )
                if cell:
                    br = cell.bounding_regions[0] if cell.bounding_regions else None
                    raw_val = cell.content.strip()
                    if raw_val.startswith("| ") or raw_val.startswith("I "):
                        raw_val = raw_val[2:]
                    raw_val = raw_val.replace("\n", " ").strip()
                    row_data[header_text] = {
                        "value": raw_val,
                        "modified": raw_val,
                        "confidence": get_cell_confidence(cell, result),
                        "polygon": br.polygon if br else None,
                        "page": br.page_number if br else None,
                    }

            if row_data and any(v["value"].strip() for v in row_data.values()):
                extracted.append(row_data)

    return extracted, result


# ==============================
# POPUP DIALOG
# ==============================
@st.dialog("Field Verification")
def show_eye_popup(field, info, pdf_path, adi_result):
    confidence = info["confidence"]
    value = info.get("modified", info["value"])
    st.markdown(f"### {field}")
    st.write("Value:")
    st.code(value)
    st.write(f"Confidence: {int(confidence * 100)}%")

    polygon = info.get("polygon")
    page_number = info.get("page")

    if polygon and page_number:
        doc = fitz.open(pdf_path)
        page = doc[page_number - 1]
        pix = page.get_pixmap(dpi=300)
        img = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGBA")

        p_width = adi_result.pages[page_number - 1].width
        p_height = adi_result.pages[page_number - 1].height
        i_width, i_height = img.size

        x_vals = polygon[0::2]
        y_vals = polygon[1::2]
        left = (min(x_vals) / p_width) * i_width
        top = (min(y_vals) / p_height) * i_height
        right = (max(x_vals) / p_width) * i_width
        bottom = (max(y_vals) / p_height) * i_height
        pad = 6
        crop_box = (
            max(0, int(left) - pad),
            max(0, int(top) - pad),
            min(i_width, int(right) + pad),
            min(i_height, int(bottom) + pad),
        )
        snippet = img.crop(crop_box)
        st.image(snippet, use_container_width=True)
        doc.close()


# ==============================
# UTILS
# ==============================
def get_val(claim, possible_keys, default=""):
    for pk in possible_keys:
        for k, v in claim.items():
            if pk.lower() in str(k).lower():
                return v["value"]
    return default


def clean_duplicate_fields(record: dict) -> dict:
    seen = set()
    cleaned = {}
    for k, v in record.items():
        key = k.strip()
        if key in seen:
            continue
        seen.add(key)
        cleaned[key] = v
    return cleaned


def save_feature_store(sheet_name, data):
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{sheet_name}_{timestamp}.json"
    path = os.path.join(FEATURE_STORE_PATH, filename)
    with open(path, "w") as f:
        json.dump(data, f, indent=2)
    return path


# ==============================
# MAIN APP
# ==============================
col_title, col_sheet_dropdown = st.columns([4, 1])
with col_title:
    st.markdown('<div class="main-title">🛡️ TPA Claims Review Portal</div>', unsafe_allow_html=True)

uploaded = st.file_uploader("Upload Loss Run Excel/CSV", type=["xlsx", "csv"])

if uploaded:
    if "tmpdir" not in st.session_state:
        st.session_state.tmpdir = tempfile.mkdtemp()

    file_ext = os.path.splitext(uploaded.name)[1]
    excel_path = os.path.join(st.session_state.tmpdir, f"input{file_ext}")

    if st.session_state.get("last_uploaded") != uploaded.name:
        with open(excel_path, "wb") as f:
            f.write(uploaded.read())
        st.session_state.last_uploaded = uploaded.name
        st.session_state.sheet_names = get_sheet_names(excel_path)
        st.session_state.sheet_cache = {}
        st.session_state.selected_idx = 0
        st.session_state.focus_field = None

    with col_sheet_dropdown:
        st.markdown("<div style='margin-top: 20px;'></div>", unsafe_allow_html=True)
        selected_sheet = st.selectbox("Sheet", st.session_state.sheet_names, label_visibility="collapsed")

    st.markdown("<hr style='border-color: #30363d; margin-top: -10px;'>", unsafe_allow_html=True)

    if selected_sheet not in st.session_state.sheet_cache:
        with st.spinner(f"Analyzing sheet '{selected_sheet}' with Azure Intelligence..."):
            pdf_path = os.path.join(st.session_state.tmpdir, f"{selected_sheet}.pdf")
            exact_headers = convert_sheet_to_pdf(excel_path, selected_sheet, pdf_path)
            data, adi_result = extract_from_adi(pdf_path, exact_headers)

            if not data:
                st.warning(f"No valid tables detected in sheet '{selected_sheet}'.")
                st.stop()

            st.session_state.sheet_cache[selected_sheet] = {
                "data": data,
                "adi_result": adi_result,
                "pdf_path": pdf_path,
            }
            st.session_state.selected_idx = 0
            st.session_state.focus_field = None

    active_cache = st.session_state.sheet_cache[selected_sheet]
    data = active_cache["data"]
    adi_result = active_cache["adi_result"]
    pdf_path = active_cache["pdf_path"]

    if st.session_state.selected_idx >= len(data):
        st.session_state.selected_idx = 0

    curr_claim = data[st.session_state.selected_idx]
    col_nav, col_main = st.columns([1.2, 3.8], gap="large")

    # ── LEFT PANEL ──
    with col_nav:
        records_container = st.container(height=600, border=False)
        with records_container:
            st.markdown(
                "<p style='color:#8b949e; font-weight:bold; font-size:12px; text-transform:uppercase;'>TPA Records</p>",
                unsafe_allow_html=True,
            )
            for i, row_data in enumerate(data):
                is_sel = "selected-card" if st.session_state.selected_idx == i else ""
                c_id = get_val(row_data, ["CLAIM_NUMBER", "Claim Number", "Claim_No", "Claim ID"], f"CLM-{10021 + i * 24}")
                c_name = get_val(row_data, ["Insured Name", "Name", "Company", "Claimant", "TPA_NAME"], "Unknown Entity")
                c_status = get_val(row_data, ["Status", "CLAIM_STATUS"], "Yet to Review" if i == 0 else "In Progress" if i == 1 else "Submitted")
                status_cls = "status-progress" if "Progress" in c_status or c_status.lower() == "open" else "status-text"

                st.markdown(f"""
                <div class="claim-card {is_sel}">
                    <div style="font-weight:bold; color:white; font-size:15px;">{c_id}</div>
                    <div style="color:#8b949e; font-size:13px; margin-top:2px;">{c_name}</div>
                    <div class="{status_cls}">{c_status}</div>
                </div>
                """, unsafe_allow_html=True)

                if st.button("Select", key=f"sel_{selected_sheet}_{i}", use_container_width=True):
                    st.session_state.selected_idx = i
                    st.session_state.focus_field = None
                    st.rerun()

    # ── RIGHT PANEL ──
    with col_main:
        head_left, head_right = st.columns([3, 1])
        curr_claim_id = get_val(
            curr_claim,
            ["CLAIM_NUMBER", "Claim Number", "Claim_No", "Claim ID"],
            f"CLM-{10021 + st.session_state.selected_idx * 24}",
        )

        with head_left:
            st.markdown(
                "<p style='color:#8b949e; font-weight:bold; font-size:12px; text-transform:uppercase;'>Review Details</p>",
                unsafe_allow_html=True,
            )
            h_name = get_val(curr_claim, ["Insured Name", "Name", "Claimant", "TPA_NAME"], "Unknown Entity")
            h_date = get_val(curr_claim, ["Loss Date", "Date", "LOSS_DATE"], "N/A")
            h_status = get_val(curr_claim, ["Status", "CLAIM_STATUS"], "Submitted")
            h_total = get_val(curr_claim, ["Total Incurred", "Incurred", "Total", "Amount", "TOTAL_INCURRED"], "$0")

            st.markdown(f"""
                <div class="mid-header-title">{curr_claim_id}</div>
                <div class="mid-header-sub">{h_name} - {h_date}</div>
                <div class="mid-header-status">{h_status}</div>
                <div class="incurred-label">Total Incurred</div>
                <div class="incurred-amount">{h_total}</div>
                <br>
                <p style='color:#8b949e; font-weight:bold; font-size:12px; text-transform:uppercase;'>Field</p>
            """, unsafe_allow_html=True)

        with head_right:
            st.markdown(
                "<p style='color:#8b949e; font-weight:bold; font-size:12px; text-transform:uppercase; text-align:right;'>Export Selection</p>",
                unsafe_allow_html=True,
            )
            b_col1, b_col2 = st.columns([1, 1])
            with b_col1:
                if st.button("☑ All", key=f"sel_all_{selected_sheet}_{curr_claim_id}", use_container_width=True):
                    for field in curr_claim.keys():
                        st.session_state[f"chk_export_{selected_sheet}_{curr_claim_id}_{field}"] = True
                    st.rerun()
            with b_col2:
                if st.button("☐ None", key=f"desel_all_{selected_sheet}_{curr_claim_id}", use_container_width=True):
                    for field in curr_claim.keys():
                        st.session_state[f"chk_export_{selected_sheet}_{curr_claim_id}_{field}"] = False
                    st.rerun()

        st.markdown("<hr style='border-color: #30363d; margin-top: -15px;'>", unsafe_allow_html=True)

        header_cols = st.columns([2, 2.6, 2.6, 0.4, 0.4, 2.2, 0.5])
        with header_cols[0]: st.markdown("**FIELD**")
        with header_cols[1]: st.markdown("**ORIGINAL VALUE**")
        with header_cols[2]: st.markdown("**MODIFIED VALUE**")
        with header_cols[3]: st.markdown(" ")
        with header_cols[4]: st.markdown(" ")
        with header_cols[5]: st.markdown("**CONFIDENCE**")
        with header_cols[6]: st.markdown(" ")

        for field, info in curr_claim.items():
            edit_key = f"edit_{selected_sheet}_{curr_claim_id}_{field}"
            widget_key = f"chk_export_{selected_sheet}_{curr_claim_id}_{field}"

            if edit_key not in st.session_state:
                st.session_state[edit_key] = False
            if widget_key not in st.session_state:
                st.session_state[widget_key] = True

            col_lbl, col_orig, col_mod, col_btn1, col_btn2, col_conf, col_chk = st.columns(
                [2, 2.6, 2.6, 0.7, 0.7, 2.2, 0.5], gap="small"
            )

            with col_lbl:
                st.markdown(
                    f"<div style='height: 40px; display: flex; align-items: center; color:#c9d1d9; font-size:12px; font-weight:bold; text-transform:uppercase;'>{field}</div>",
                    unsafe_allow_html=True,
                )
            with col_orig:
                st.text_input(
                    "orig",
                    value=info["value"],
                    key=f"orig_{selected_sheet}_{curr_claim_id}_{field}",
                    label_visibility="collapsed",
                    disabled=True,
                )
            with col_mod:
                edit_enabled = st.session_state[edit_key]
                mod_key = f"mod_{selected_sheet}_{curr_claim_id}_{field}"
                if mod_key not in st.session_state:
                    st.session_state[mod_key] = info.get("modified", info["value"])
                new_val = st.text_input("mod", key=mod_key, label_visibility="collapsed", disabled=not edit_enabled)
                st.session_state.sheet_cache[selected_sheet]["data"][st.session_state.selected_idx][field]["modified"] = new_val

            with col_btn1:
                if st.button("👁", key=f"view_{selected_sheet}_{curr_claim_id}_{field}", use_container_width=True):
                    show_eye_popup(field, info, pdf_path, adi_result)

            with col_btn2:
                conf = info["confidence"]
                if conf > 0.98:
                    st.button("🔒", key=f"locked_{selected_sheet}_{curr_claim_id}_{field}", disabled=True, use_container_width=True)
                else:
                    if st.button("✏", key=f"toggle_edit_{selected_sheet}_{curr_claim_id}_{field}", use_container_width=True):
                        st.session_state[edit_key] = not st.session_state[edit_key]
                        st.rerun()

            with col_conf:
                conf = info["confidence"]
                color = "#3fb950" if conf >= 0.85 else "#d29922" if conf >= 0.70 else "#f85149"
                st.markdown(f"""
                    <div style="display: flex; align-items: center; gap: 10px; height: 40px;">
                        <div style="flex-grow: 1; height: 6px; background: #30363d; border-radius: 3px;">
                            <div style="width:{conf*100}%; height:100%; background:{color}; box-shadow: 0 0 5px {color}; border-radius: 3px;"></div>
                        </div>
                        <div style="font-size: 12px; color: #8b949e; width: 35px; text-align: right;">{int(conf*100)}%</div>
                    </div>
                """, unsafe_allow_html=True)

            with col_chk:
                st.markdown("<div style='height: 8px;'></div>", unsafe_allow_html=True)
                st.checkbox("", key=widget_key, label_visibility="collapsed")

        st.markdown("<hr style='border-color: #30363d;'>", unsafe_allow_html=True)

        btn_col1, btn_col2 = st.columns([4, 1.5])
        with btn_col2:
            if st.button(f"☑ Export Sheet '{selected_sheet}' to JSON", type="primary", use_container_width=True):
                export_data = {}
                for i, row in enumerate(data):
                    c_id = get_val(row, ["CLAIM_NUMBER", "Claim Number", "Claim_No"], f"CLM-{10021 + i * 24}")
                    export_data[c_id] = {}
                    for fld, inf in row.items():
                        if st.session_state.get(f"chk_export_{selected_sheet}_{c_id}_{fld}", True):
                            val = inf.get("modified") if inf.get("modified") != inf.get("value") else inf.get("value")
                            export_data[c_id][fld] = val
                    export_data[c_id] = clean_duplicate_fields(export_data[c_id])

                saved_path = save_feature_store(selected_sheet, export_data)
                json_str = json.dumps(export_data, indent=2)
                st.success(f"Saved to Feature Store → {saved_path}")
                st.download_button(
                    "📥 Download JSON",
                    data=json_str,
                    file_name=f"{selected_sheet}_validated.json",
                    mime="application/json",
                    use_container_width=True,
                )

    # ── BOUNDING BOX VIEWER ──
    if st.session_state.focus_field:
        st.markdown("<br><br>", unsafe_allow_html=True)
        field = st.session_state.focus_field
        info = curr_claim[field]
        polygon, page_number = info.get("polygon"), info.get("page")

        if polygon and page_number:
            with st.spinner("Loading document spotlight..."):
                doc = fitz.open(pdf_path)
                page = doc[page_number - 1]
                pix = page.get_pixmap(dpi=300)
                img = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGBA")

                p_width = adi_result.pages[page_number - 1].width
                p_height = adi_result.pages[page_number - 1].height
                i_width, i_height = img.size
                x_vals, y_vals = polygon[0::2], polygon[1::2]
                left = (min(x_vals) / p_width) * i_width
                top = (min(y_vals) / p_height) * i_height
                right = (max(x_vals) / p_width) * i_width
                bottom = (max(y_vals) / p_height) * i_height
                pad = 4
                crop_box = (
                    max(0, int(left) - pad),
                    max(0, int(top) - pad),
                    min(i_width, int(right) + pad),
                    min(i_height, int(bottom) + pad),
                )

                overlay = Image.new("RGBA", img.size, (13, 17, 23, 220))
                darkened = Image.alpha_composite(img, overlay)
                darkened.paste(img.crop(crop_box), crop_box)
                ImageDraw.Draw(darkened).rectangle(crop_box, outline="#58a6ff", width=5)

                st.markdown(
                    f"<p style='color:white; font-size:16px;'>Spotlight: <span style='color:#58a6ff; font-weight:bold;'>{field}</span> (Page {page_number})</p>",
                    unsafe_allow_html=True,
                )
                st.image(darkened, use_container_width=True)
                doc.close()
        else:
            st.warning(f"No spatial coordinates found for {field}.")
