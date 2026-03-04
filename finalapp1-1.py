import streamlit as st
import os
import json
import tempfile
import csv
import datetime
import openpyxl
from openpyxl.utils import get_column_letter

# ==============================
# FEATURE STORE
# ==============================
FEATURE_STORE_PATH="feature_store/claims_json"
os.makedirs(FEATURE_STORE_PATH,exist_ok=True)

# ==============================
# PAGE CONFIG
# ==============================
st.set_page_config(layout="wide",page_title="TPA Claims Review Portal")

# ==============================
# STYLING
# ==============================
st.markdown("""
<style>

.stApp{background:#0d1117;color:#c9d1d9;}

.main-title{
font-size:26px;
font-weight:600;
padding:10px 0;
border-bottom:1px solid #30363d;
margin-bottom:20px;
color:white;
}

.claim-card{
background:#161b22;
border:1px solid #30363d;
border-radius:8px;
padding:15px;
margin-bottom:10px;
}

.selected-card{
border-left:4px solid #58a6ff;
background:#1c2128;
}

.status-text{font-size:12px;color:#3fb950;}

.cell-index{
font-size:11px;
color:#8b949e;
}

</style>
""",unsafe_allow_html=True)

# ==============================
# METADATA EXTRACTION
# ==============================
def extract_metadata(rows):

    report_name=rows[0][0] if rows else ""

    # Use ordered list to preserve original order of metadata fields
    summary=[]

    if len(rows)>1:

        for cell in rows[1]:

            if ":" in str(cell):

                k,v=str(cell).split(":",1)

                summary.append((k.strip(), v.strip()))

    return report_name,summary


# ==============================
# TEXT SANITIZATION
# ==============================
def sanitize_text_for_markdown(text):
    """Escape # characters to prevent markdown interpretation"""
    if text and isinstance(text, str):
        # Replace # with HTML entity to prevent markdown interpretation
        return text.replace("#", "&#35;")
    return text


# ==============================
# SHEET NAMES
# ==============================
def get_sheet_names(file_path):

    ext=os.path.splitext(file_path)[1].lower()

    if ext==".csv":
        return["Sheet1"]

    wb=openpyxl.load_workbook(file_path,read_only=True)

    names=wb.sheetnames

    wb.close()

    return names


# ==============================
# EXTRACT FROM EXCEL
# ==============================
def extract_from_excel(file_path, sheet_name):

    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
    else:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[sheet_name]

        rows = []

        for r in ws.iter_rows():
            row = []

            for c in r:

                val = c.value

                if val is None:
                    row.append("")

                elif isinstance(val, datetime.datetime):
                    row.append(val.strftime("%m-%d-%Y"))

                elif isinstance(val, datetime.date):
                    row.append(val.strftime("%m-%d-%Y"))

                else:
                    row.append(str(val).strip())

            rows.append(row)

        wb.close()

    report_name, summary = extract_metadata(rows)

    # ==============================
    # FIND HEADER ROW
    # ==============================

    header_row_idx = None

    claim_headers = [
        "claim id","claim number","claim no","claim#",
        "claim ref","claim reference","file number","file no"
    ]

    for i, row in enumerate(rows[:30]):

        for cell in row:

            txt = str(cell).lower().strip()

            if any(x in txt for x in claim_headers):
                header_row_idx = i
                break

        if header_row_idx is not None:
            break

    # fallback logic
    if header_row_idx is None:

        for i, row in enumerate(rows[:30]):

            non_empty = sum(1 for c in row if c)

            if non_empty >= 3:

                row_str = " ".join([str(c).lower() for c in row if c])

                if "claim" in row_str or "date" in row_str:
                    header_row_idx = i
                    break

    if header_row_idx is None:
        header_row_idx = 2

    headers = rows[header_row_idx]

    headers_list = [
        h for h in headers if h and h.lower() not in ["claims","#claims"]
    ]

    extracted = []
    totals = {}

    for r_idx, r in enumerate(rows[header_row_idx+1:], start=header_row_idx+2):

        row_join = " ".join(r).lower()

        if row_join.strip().startswith("total"):

            for c_idx, val in enumerate(r):

                if val:
                    totals[headers[c_idx]] = val

            continue

        if not any(r):
            continue

        row_data = {}

        for c_idx, val in enumerate(r):

            if c_idx >= len(headers):
                continue

            header = headers[c_idx]

            if header.lower() in ["claims","#claims"]:
                continue

            col_letter = get_column_letter(c_idx+1)

            cell_index = f"{col_letter}{r_idx}"

            if val == "":
                val = "0"

            row_data[header] = {
                "value": val,
                "modified": val,
                "cell": cell_index
            }

        extracted.append(row_data)

    return report_name, summary, totals, extracted, headers_list

# ==============================
# UTILITIES
# ==============================
def detect_claim_id(row, index=None):

    keys = [
        "claim id","claim number","claim no",
        "claim#","claim ref","claim reference",
        "file number","record id"
    ]

    for k, v in row.items():

        name = k.lower().replace("_"," ").strip()

        if any(x in name for x in keys):

            val = v.get("modified") or v.get("value")

            if val:
                return str(val)

    if index is not None:
        return str(index + 1)

    return ""

def clean_duplicate_fields(record):

    seen=set()
    out={}

    for k,v in record.items():

        if k.strip() not in seen:

            seen.add(k.strip())

            out[k.strip()]=v

    return out


def save_feature_store(sheet_name,data):

    ts=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    path=os.path.join(FEATURE_STORE_PATH,f"{sheet_name}_{ts}.json")

    with open(path,"w") as f:

        json.dump(data,f,indent=2)

    return path


# ==============================
# EXCEL REGENERATION
# ==============================
def regenerate_excel(report,summary,totals,claims,output,headers_list):

    wb=openpyxl.Workbook()

    ws=wb.active

    ws.append([report])

    # Filter out Report Name from summary (summary is now a list of tuples)
    filtered_summary = [(k, v) for k, v in summary if k.lower().replace(" ", "") != "reportname"]

    meta_line=" | ".join([f"{k}: {v}" for k,v in filtered_summary])

    ws.append([meta_line])

    ws.append([])

    # Use the original headers order from headers_list
    headers = headers_list if headers_list else list(next(iter(claims.values())).keys())

    ws.append(headers)

    # Use the same order as headers_list for each claim
    for claim in claims.values():

        row=[claim.get(h,"") for h in headers]

        ws.append(row)

    if totals:

        total_row=[]

        for h in headers:

            total_row.append(totals.get(h,""))

        ws.append(total_row)

    wb.save(output)


# ==============================
# EYE POPUP
# ==============================
@st.dialog("Field Verification")
def show_eye_popup(field,info):

    st.markdown(f"### {field}")

    value=info.get("modified",info["value"])

    st.code(value if value else "(empty)")


# ==============================
# MAIN APP
# ==============================
col_title,col_sheet_dropdown=st.columns([4,1])

with col_title:

    st.markdown('<div class="main-title">🛡️ TPA Claims Review Portal</div>',unsafe_allow_html=True)

uploaded=st.file_uploader("Upload Loss Run Excel/CSV",type=["xlsx","csv"])

if uploaded:

    if "tmpdir" not in st.session_state:

        st.session_state.tmpdir=tempfile.mkdtemp()

    file_ext=os.path.splitext(uploaded.name)[1]

    excel_path=os.path.join(st.session_state.tmpdir,f"input{file_ext}")

    if st.session_state.get("last_uploaded")!=uploaded.name:

        with open(excel_path,"wb") as f:

            f.write(uploaded.read())

        st.session_state.last_uploaded=uploaded.name

        st.session_state.sheet_names=get_sheet_names(excel_path)

        st.session_state.sheet_cache={}

        st.session_state.selected_idx=0

        # Initialize processed_data for tracking modifications
        st.session_state.processed_data={}

    with col_sheet_dropdown:

        selected_sheet=st.selectbox("Sheet",st.session_state.sheet_names,label_visibility="collapsed")

    if selected_sheet not in st.session_state.sheet_cache:

        report,summary,totals,data,headers_list=extract_from_excel(excel_path,selected_sheet)

        st.session_state.sheet_cache[selected_sheet]={
            "data":data,
            "report":report,
            "summary":summary,
            "totals":totals,
            "headers_list":headers_list
        }

    cache=st.session_state.sheet_cache[selected_sheet]

    report_name=cache["report"]

    summary=cache["summary"]

    totals=cache["totals"]

    data=cache["data"]

    headers_list=cache.get("headers_list", [])

    # Ensure data is properly synced - get the actual data reference from cache
    # This ensures modifications persist across reruns
    if "processed_data" not in st.session_state:
        st.session_state.processed_data = {}

    # Store the current data in session state for persistence
    st.session_state.processed_data[selected_sheet] = data

    curr_claim=data[st.session_state.selected_idx]

    col_nav,col_main=st.columns([1.2,3.8])

# ==============================
# LEFT PANEL
# ==============================
    with col_nav:

        st.markdown("**TPA RECORDS**")

        scroll_container=st.container(height=650)

        with scroll_container:

            for i,row_data in enumerate(data):

                is_sel="selected-card" if st.session_state.selected_idx==i else ""

                c_id=detect_claim_id(row_data,i)

                st.markdown(f"""
                <div class="claim-card {is_sel}">
                <div style="font-weight:bold;color:white;">{c_id}</div>
                <div class="status-text">Claim Record</div>
                </div>
                """,unsafe_allow_html=True)

                if st.button("Select",key=f"sel_{i}",use_container_width=True):

                    st.session_state.selected_idx=i

                    st.rerun()

# ==============================
# RIGHT PANEL
# ==============================
    with col_main:

        st.markdown("### Report Name")

        st.text_input("",report_name,disabled=True)

        st.markdown("### Summary")

        # Summary is now a list of tuples, iterate directly
        for k,v in summary:

            # Skip "Report Name" in summary display
            if k.lower().replace(" ", "") == "reportname":
                continue

            st.text_input(k,v,disabled=True)

        st.markdown("---")

        curr_claim_id=detect_claim_id(curr_claim)

        st.markdown(f"### {curr_claim_id}")

        header=st.columns([2,3,3,0.6,0.6,0.5])

        header[0].markdown("**FIELD**")

        header[1].markdown("**EXTRACTED**")

        header[2].markdown("**MODIFIED**")

        for field,info in curr_claim.items():

            edit_key=f"edit_{curr_claim_id}_{field}"

            mod_key=f"mod_{curr_claim_id}_{field}"

            chk_key=f"chk_{curr_claim_id}_{field}"

            if edit_key not in st.session_state:
                st.session_state[edit_key]=False

            if mod_key not in st.session_state:
                st.session_state[mod_key]=info["value"]

            cols=st.columns([2,3,3,0.6,0.6,0.5])

            with cols[0]:

                # Sanitize field name to prevent # being interpreted as markdown header
                safe_field = sanitize_text_for_markdown(field)

                st.markdown(f"""
                {safe_field}
                <div class='cell-index'>Excel: {info['cell']}</div>
                """,unsafe_allow_html=True)

            with cols[1]:

                st.text_input("",value=info["value"],disabled=True,key=f"orig_{curr_claim_id}_{field}")

            with cols[2]:

                # Use session state value directly, no default value parameter to avoid warning
                nv=st.text_input("",key=mod_key,disabled=not st.session_state[edit_key])

                # Explicitly update the info object with the session state value
                info["modified"] = st.session_state.get(mod_key, info["value"])

            with cols[3]:

                if st.button("👁",key=f"eye_{curr_claim_id}_{field}"):

                    show_eye_popup(field,info)

            with cols[4]:

                if st.button("✏",key=f"edit_btn_{curr_claim_id}_{field}"):

                    st.session_state[edit_key]=not st.session_state[edit_key]

                    st.rerun()

            with cols[5]:

                st.checkbox("",key=chk_key,value=True)

        # BUILD EXPORT DATA
        # Use processed_data from session state to ensure modifications persist
        export_data = {}

        # Get data from session state to ensure we have the latest modifications
        export_source_data = st.session_state.processed_data.get(selected_sheet, data)

        for i,row in enumerate(export_source_data):

            cid=detect_claim_id(row,i)

            rec={}

            for fld,inf in row.items():

                rec[fld]=inf["modified"]

            export_data[cid]=clean_duplicate_fields(rec)

        if st.button("Export JSON"):

            final_json={
                "report_name":report_name,
                "summary":summary,
                "totals":totals,
                "claims":export_data
            }

            saved=save_feature_store(selected_sheet,final_json)

            json_str=json.dumps(final_json,indent=2)

            st.success(f"Saved → {saved}")

            st.download_button(
                "Download JSON",
                data=json_str,
                file_name="claims.json",
                mime="application/json"
            )

        if st.button("Regenerate Excel"):

            regen_path=os.path.join(st.session_state.tmpdir,"reconstructed.xlsx")

            regenerate_excel(report_name,summary,totals,export_data,regen_path,headers_list)

            with open(regen_path,"rb") as f:

                st.download_button(
                    "Download Excel",
                    data=f,
                    file_name="reconstructed_loss_run.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )