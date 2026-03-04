import streamlit as st
import os
import json
import tempfile
import csv
import datetime
import re
import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw

# ==============================
# FEATURE STORE
# ==============================
FEATURE_STORE_PATH = "feature_store/claims_json"
os.makedirs(FEATURE_STORE_PATH, exist_ok=True)

# ==============================
# PAGE CONFIG
# ==============================
st.set_page_config(layout="wide", page_title="TPA Claims Review Portal")
if "focus_field" not in st.session_state:
    st.session_state.focus_field = None

# ==============================
# STYLING
# ==============================
st.markdown("""
<style>
    .stApp { background-color: #0d1117; color: #c9d1d9; }
    .main-title {
        font-size: 26px; font-weight: 600; padding: 10px 0;
        border-bottom: 1px solid #30363d; margin-bottom: 20px; color: white;
        text-shadow: 0 0 10px rgba(88,166,255,0.7);
    }
    .claim-card {
        background: #161b22; border: 1px solid #30363d; border-radius: 8px;
        padding: 15px; margin-bottom: 10px; cursor: pointer;
        box-shadow: 0 0 0 transparent; transition: all .25s ease;
    }
    .claim-card:hover {
        border-color: #58a6ff;
        box-shadow: 0 0 12px rgba(88,166,255,0.6);
        transform: translateY(-2px);
    }
    .selected-card { border-left: 4px solid #58a6ff; background: #1c2128; box-shadow: 0 0 16px rgba(88,166,255,0.8); }
    .status-text     { font-size: 12px; color: #3fb950; margin-top: 5px; }
    .status-progress { font-size: 12px; color: #d29922; margin-top: 5px; }
    .mid-header-title  { font-size: 26px; font-weight: bold; color: white; margin-bottom: 0px; }
    .mid-header-sub    { font-size: 15px; color: #8b949e; margin-top: 5px; margin-bottom: 5px; }
    .mid-header-status { font-size: 13px; color: #3fb950; margin-bottom: 15px; }
    .incurred-label    { font-size: 14px; color: #8b949e; margin-bottom: 0px; }
    .incurred-amount   { font-size: 26px; font-weight: bold; color: #3fb950; margin-top: 0px; margin-bottom: 20px; }
    div[data-baseweb="input"],
    div[data-baseweb="base-input"],
    div[data-baseweb="select"] {
        background-color: #161b22 !important;
        border: 1px solid #30363d !important;
        border-radius: 6px !important;
    }
    div[data-baseweb="input"] input {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        background-color: transparent !important;
        font-size: 15px !important;
        padding: 8px 12px !important;
    }
    div[data-baseweb="input"]:has(input:disabled),
    div[data-baseweb="base-input"]:has(input:disabled) {
        background-color: transparent !important;
        border: none !important;
    }
    div[data-baseweb="input"] input:disabled {
        color: #e6edf3 !important;
        -webkit-text-fill-color: #e6edf3 !important;
        cursor: default !important;
        padding-left: 0px !important;
    }
    div[data-testid="stButton"] button {
        background-color: transparent !important;
        color: #8b949e !important;
        border: 1px solid #30363d !important;
        border-radius: 6px !important;
        padding: 2px 8px !important;
        transition: 0.2s;
    }
    div[data-testid="stButton"] button:hover {
        border-color: #58a6ff !important;
        color: #58a6ff !important;
        background-color: #1c2128 !important;
    }
    div[data-testid="stButton"] button:disabled { opacity: 0.3 !important; }
    div[role="dialog"] {
        background-color: #0d1117 !important;
        border: 1px solid #30363d !important;
        border-radius: 10px !important;
    }
    div[role="dialog"] * { color: #c9d1d9 !important; }
    div[role="dialog"] button {
        background-color: transparent !important;
        border: 1px solid #30363d !important;
        color: #8b949e !important;
    }
    div[role="dialog"] button:hover {
        border-color: #58a6ff !important;
        color: #58a6ff !important;
        background-color: #1c2128 !important;
    }
    .left-scroll-container { height: calc(100vh - 140px); overflow-y: auto; padding-right: 6px; }
</style>
""", unsafe_allow_html=True)


# ==============================
# SHEET NAMES
# ==============================
def get_sheet_names(file_path: str) -> list:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return ["Sheet1"]
    wb    = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def calculate_confidence(header, value):
    if not value:
        return 0.5
    header = header.lower()
    value  = str(value)
    if "date" in header:
        try:
            datetime.datetime.strptime(value[:10], "%m-%d-%Y")
            return 0.97
        except:
            return 0.65
    if "claim" in header:
        if re.match(r"[A-Z0-9\-]+", value):
            return 0.95
    if "amount" in header or "incurred" in header:
        try:
            float(value.replace(",", ""))
            return 0.98
        except:
            return 0.70
    return 0.85


# ==============================
# CLASSIFICATION
# ==============================
def classify_sheet(rows):
    text = " ".join(
        str(cell).lower()
        for row in rows[:20]
        for cell in row
        if cell
    )
    if "line of business" in text:
        return "SUMMARY"
    has_claim = any(x in text for x in [
        "claim number", "claim no", "claim #", "claim id",
        "claim ref", "claimant", "file number", "file no"
    ])
    has_loss = any(x in text for x in [
        "loss date", "date of loss", "loss dt", "accident date",
        "occurrence date", "incident date"
    ])
    has_financial = any(x in text for x in [
        "incurred", "paid", "reserve", "outstanding",
        "total paid", "total incurred", "indemnity", "expense"
    ])
    if has_claim and (has_loss or has_financial):
        return "LOSS_RUN"
    if "policy" in text and ("claim" in text or "incurred" in text):
        return "COMMERCIAL_LOSS_RUN"
    if has_claim:
        return "LOSS_RUN"
    return "UNKNOWN"
def extract_report_header(rows):

    header_lines = []

    if not rows:
        return ""

    max_scan = min(10, len(rows))   # prevent overflow

    for i in range(max_scan):

        row = rows[i]

        text = " ".join(str(x) for x in row if x)

        text_lower = text.lower()

        if any(x in text_lower for x in [
            "claim id","claim number","claimant",
            "loss date","date reported","total incurred"
        ]):
            break

        if len(text.strip()) > 10:
            header_lines.append(text)

    return " | ".join(header_lines)

def extract_from_excel(file_path, sheet_name):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
    else:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        with open("debug_openpyxl.json", "w") as f:
            json.dump(rows[:20], f, indent=2, default=str)
        wb.close()
    if not rows:
        return [], "UNKNOWN",""
    sheet_type = classify_sheet(rows)

    report_header = extract_report_header(rows)

    data, sheet_type = parse_rows(sheet_type, rows)

    return data, sheet_type, report_header
    # print(f"Detected sheet type: {sheet_type}")
    # return parse_rows(sheet_type, rows)


def parse_rows(sheet_type, rows):
    # ── SUMMARY ──
    if sheet_type == "SUMMARY":
        header_row_index = None
        for i, row in enumerate(rows[:20]):
            row_text = " ".join([str(c).lower() for c in row if c])
            if "sheet" in row_text and "line of business" in row_text:
                header_row_index = i
                break
        if header_row_index is None:
            return [], sheet_type
        headers = [
            str(h).strip() if h is not None else f"Column_{i}"
            for i, h in enumerate(rows[header_row_index])
        ]
        extracted = []
        for r_idx, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
            if not any(row):
                continue
            row_data = {}
            for c_idx, value in enumerate(row, start=1):
                if c_idx - 1 >= len(headers):
                    continue
                header    = headers[c_idx - 1]
                clean_val = str(value).strip() if value is not None else ""
                conf      = calculate_confidence(header, clean_val)
                row_data[header] = {
                    "value": clean_val, "modified": clean_val,
                    "confidence": conf, "excel_row": r_idx, "excel_col": c_idx,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    # ── LOSS RUN ──
    header_row_index = None
    for i, row in enumerate(rows[:20]):
        row_text = " ".join([str(c).lower() for c in row if c])
        if "claim" in row_text and (
            "date" in row_text or "incurred" in row_text or "paid" in row_text
        ):
            header_row_index = i
            break

    if header_row_index is None:
        return [], sheet_type

    headers = [
        str(h).strip() if h is not None else f"Column_{i}"
        for i, h in enumerate(rows[header_row_index])
    ]
    extracted = []
    for r_idx, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
        if not any(row):
            continue
        if any(str(cell).lower().strip() == "totals" for cell in row if cell):
            break
        row_data = {}
        for c_idx, value in enumerate(row, start=1):
            if c_idx - 1 >= len(headers):
                continue
            header    = headers[c_idx - 1]
            clean_val = str(value).strip() if value is not None else ""
            conf      = calculate_confidence(header, clean_val)
            row_data[header] = {
                "value": clean_val, "modified": clean_val,
                "confidence": conf, "excel_row": r_idx, "excel_col": c_idx,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


# ======================================================
# PURE PYTHON EXCEL RENDERER  (no LibreOffice needed)
# ======================================================

# Standard Office theme colors (index 0-9)
_THEME_COLORS = {
    0: "FFFFFF", 1: "000000", 2: "EEECE1", 3: "1F497D",
    4: "4F81BD", 5: "C0504D", 6: "9BBB59", 7: "8064A2",
    8: "4BACC6", 9: "F79646",
}


def _resolve_color(color_obj, default="FFFFFF") -> str:
    """Convert an openpyxl Color object → 6-char hex RGB string."""
    if color_obj is None:
        return default
    t = color_obj.type

    if t == "rgb":
        rgb = color_obj.rgb or ""
        # AARRGGBB — strip alpha; skip fully-transparent black (00000000)
        if len(rgb) == 8 and rgb not in ("00000000", "FF000000"):
            return rgb[2:]   # drop AA prefix → RRGGBB
        if len(rgb) == 6:
            return rgb
        return default

    if t == "theme":
        base = _THEME_COLORS.get(color_obj.theme, default)
        tint = color_obj.tint or 0.0
        if tint != 0.0:
            r, g, b = int(base[0:2], 16), int(base[2:4], 16), int(base[4:6], 16)
            if tint > 0:
                r = int(r + (255 - r) * tint)
                g = int(g + (255 - g) * tint)
                b = int(b + (255 - b) * tint)
            else:
                r = int(r * (1 + tint))
                g = int(g * (1 + tint))
                b = int(b * (1 + tint))
            return f"{max(0,min(255,r)):02X}{max(0,min(255,g)):02X}{max(0,min(255,b)):02X}"
        return base

    if t == "indexed":
        # Indexed colors — common ones
        indexed_map = {
            0: "000000", 1: "FFFFFF", 2: "FF0000", 3: "00FF00",
            4: "0000FF", 5: "FFFF00", 6: "FF00FF", 7: "00FFFF",
            64: "000000", 65: "FFFFFF",
        }
        return indexed_map.get(color_obj.indexed, default)

    return default


def _col_px(ws, c: int, scale: float = 1.0) -> int:
    letter = get_column_letter(c)
    cd = ws.column_dimensions.get(letter)
    w  = cd.width if (cd and cd.width and cd.width > 0) else 8.43
    return max(20, int(w * 8 * scale))


def _row_px(ws, r: int, scale: float = 1.0) -> int:
    rd = ws.row_dimensions.get(r)
    h  = rd.height if (rd and rd.height and rd.height > 0) else 15.0
    return max(14, int(h * 1.5 * scale))


def render_excel_sheet(excel_path: str, sheet_name: str,
                        scale: float = 1.0) -> tuple[Image.Image, list, list]:
    """
    Render an Excel worksheet to a PIL Image using openpyxl styles.
    Returns (image, col_starts, row_starts) where *_starts are pixel positions.
    No external tools required — pure Python / Pillow.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    # ── Pixel grid ──────────────────────────────────────────────────────
    col_starts = [0]
    for c in range(1, max_col + 1):
        col_starts.append(col_starts[-1] + _col_px(ws, c, scale))

    row_starts = [0]
    for r in range(1, max_row + 1):
        row_starts.append(row_starts[-1] + _row_px(ws, r, scale))

    img_w = col_starts[-1]
    img_h = row_starts[-1]

    img  = Image.new("RGB", (img_w, img_h), "white")
    draw = ImageDraw.Draw(img, "RGBA")

    # ── Merged cell map ─────────────────────────────────────────────────
    merged_master: dict[tuple, tuple] = {}   # (r,c) → (min_r,min_c,max_r,max_c)
    for mr in ws.merged_cells.ranges:
        mn_r, mn_c, mx_r, mx_c = mr.bounds
        for rr in range(mn_r, mx_r + 1):
            for cc in range(mn_c, mx_c + 1):
                merged_master[(rr, cc)] = (mn_r, mn_c, mx_r, mx_c)

    drawn_merges: set = set()

    # ── Draw each cell ───────────────────────────────────────────────────
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):

            merge_info = merged_master.get((r, c))

            if merge_info:
                mn_r, mn_c, mx_r, mx_c = merge_info
                if (mn_r, mn_c) in drawn_merges:
                    continue          # already painted
                drawn_merges.add((mn_r, mn_c))
                x1 = col_starts[mn_c - 1];  y1 = row_starts[mn_r - 1]
                x2 = col_starts[mx_c];       y2 = row_starts[mx_r]
                cell = ws.cell(mn_r, mn_c)   # master carries the style
            else:
                x1 = col_starts[c - 1];  y1 = row_starts[r - 1]
                x2 = col_starts[c];       y2 = row_starts[r]
                cell = ws.cell(r, c)

            # ── Background fill ────────────────────────────────────────
            bg_hex = "FFFFFF"
            if cell.fill and cell.fill.fill_type == "solid":
                bg_hex = _resolve_color(cell.fill.fgColor, "FFFFFF")

            draw.rectangle([x1, y1, x2 - 1, y2 - 1],
                           fill=f"#{bg_hex}", outline="#CCCCCC", width=1)

            # ── Cell text ──────────────────────────────────────────────
            val = cell.value
            if val is not None:
                txt_color = "#000000"
                if cell.font and cell.font.color:
                    fc = _resolve_color(cell.font.color, "000000")
                    # Avoid white-on-white
                    if fc.upper() != bg_hex.upper():
                        txt_color = f"#{fc}"

                bold = bool(cell.font and cell.font.bold)
                text = str(val)

                # Truncate to fit cell width
                cell_w   = x2 - x1
                ch_w     = 7 if not bold else 8
                max_chars = max(1, (cell_w - 8) // ch_w)
                if len(text) > max_chars:
                    text = text[:max_chars - 1] + "…"

                draw.text((x1 + 4, y1 + 4), text, fill=txt_color)

    wb.close()
    return img, col_starts, row_starts


def get_cell_pixel_bbox(col_starts: list, row_starts: list,
                         target_row: int, target_col: int) -> tuple[int, int, int, int]:
    """Return (x1,y1,x2,y2) pixel bbox for a cell given precomputed start arrays."""
    c = min(target_col, len(col_starts) - 1)
    r = min(target_row, len(row_starts) - 1)
    return col_starts[c - 1], row_starts[r - 1], col_starts[c], row_starts[r]


def crop_context(img: Image.Image, x1, y1, x2, y2,
                 pad_x=220, pad_y=160) -> tuple[Image.Image, int, int, int, int]:
    """Crop a padded window around the target cell. Returns (img, nx1,ny1,nx2,ny2)."""
    iw, ih = img.size
    cx1 = max(0, x1 - pad_x);  cy1 = max(0, y1 - pad_y)
    cx2 = min(iw, x2 + pad_x); cy2 = min(ih, y2 + pad_y)
    cropped = img.crop((cx1, cy1, cx2, cy2))
    return cropped, x1 - cx1, y1 - cy1, x2 - cx1, y2 - cy1


# # ==============================
# # EYE POPUP
# # ==============================
# @st.dialog("Field Verification", width="large")
# def show_eye_popup(field, info, excel_path, sheet_name):
#     st.markdown(f"### 📍 {field}")

#     value = info.get("modified", info["value"])
#     col_a, col_b = st.columns([1, 1])
#     with col_a:
#         st.markdown("**Extracted Value**")
#         st.code(value if value else "(empty)")
#     with col_b:
#         conf  = info.get("confidence", 0.99)
#         color = "#3fb950" if conf >= 0.85 else "#d29922" if conf >= 0.70 else "#f85149"
#         r_lbl = info.get("excel_row", "?")
#         c_lbl = info.get("excel_col", "?")
#         st.markdown(f"""
#             <div style="padding:10px 0;">
#                 <div style="margin-bottom:6px;">
#                     <span style="color:#8b949e;">Confidence: </span>
#                     <span style="color:{color}; font-weight:bold; font-size:18px;">{int(conf*100)}%</span>
#                     <span style="color:#8b949e; font-size:12px; margin-left:12px;">
#                         Row {r_lbl} · Col {c_lbl}
#                     </span>
#                 </div>
#                 <div style="height:8px; background:#30363d; border-radius:4px;">
#                     <div style="width:{conf*100}%; height:100%; background:{color};
#                                 box-shadow:0 0 6px {color}; border-radius:4px;"></div>
#                 </div>
#             </div>
#         """, unsafe_allow_html=True)

#     target_row = info.get("excel_row")
#     target_col = info.get("excel_col")

#     if not target_row or not target_col:
#         st.warning("No cell location recorded for this field.")
#         return

#     ext = os.path.splitext(excel_path)[1].lower()
#     if ext == ".csv":
#         st.info("Cell preview is not available for CSV files.")
#         return

#     st.markdown("---")
#     st.markdown("**📊 Excel Cell Location**")

#     # Cache the rendered sheet image so it's only built once per sheet
#     cache_key = f"_rendered_{excel_path}_{sheet_name}"
#     with st.spinner("Rendering sheet…"):
#         if cache_key not in st.session_state:
#             rendered_img, col_starts, row_starts = render_excel_sheet(
#                 excel_path, sheet_name, scale=1.0
#             )
#             st.session_state[cache_key] = (rendered_img, col_starts, row_starts)
#         else:
#             rendered_img, col_starts, row_starts = st.session_state[cache_key]

#     try:
#         # Work on a copy so we don't corrupt the cache
#         img = rendered_img.copy()
#         draw = ImageDraw.Draw(img, "RGBA")

#         x1, y1, x2, y2 = get_cell_pixel_bbox(col_starts, row_starts,
#                                                target_row, target_col)

#         # Semi-transparent red highlight fill
        


#         # Crop to context window
#         cropped, nx1, ny1, nx2, ny2 = crop_context(img, x1, y1, x2, y2,
#                                                      pad_x=250, pad_y=180)

#         col_letter = get_column_letter(target_col)
#         st.image(
#             cropped,
#             use_container_width=True,
#             caption=f"Cell {col_letter}{target_row}  ·  Row {target_row}, Col {target_col}  ·  Value: {value or '(empty)'}"
#         )

#     except Exception as e:
#         st.error(f"Rendering error: {e}")
#         import traceback
#         st.code(traceback.format_exc())


# ==============================
# UTILS
# ==============================
def get_val(claim: dict, keys: list, default: str = "") -> str:
    for pk in keys:
        for k, v in claim.items():
            if pk.lower() in str(k).lower():
                return v["value"] or default
    return default


def detect_claim_id(row, index=None):
    keys = [
        "claim id", "claim_id", "claimid",
        "claim number", "claim no", "claim #",
        "claim ref", "claim reference",
        "file number", "record id"
    ]
    for k, v in row.items():
        name = str(k).lower().replace("_", " ").strip()
        if any(x in name for x in keys):
            val = v.get("modified") or v.get("value")
            if val and str(val).strip():
                return str(val)
    if index is not None:
        return str(index + 1)
    return ""


def clean_duplicate_fields(record: dict) -> dict:
    seen, out = set(), {}
    for k, v in record.items():
        if k.strip() not in seen:
            seen.add(k.strip())
            out[k.strip()] = v
    return out


def save_feature_store(sheet_name: str, data: dict) -> str:
    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(FEATURE_STORE_PATH, f"{sheet_name}_{ts}.json")
    with open(path, "w") as f:
        json.dump(data, f, indent=2)
    return path


# ==============================
# MAIN APP
# ==============================
col_title, col_sheet_dropdown = st.columns([4, 1])
with col_title:
    st.markdown('<div class="main-title">🛡️ TPA Claims Review Portal</div>', unsafe_allow_html=True)

uploaded = st.file_uploader("Upload Loss Run Excel/CSV", type=["xlsx", "csv"])

if uploaded:
    if "tmpdir" not in st.session_state:
        st.session_state.tmpdir = tempfile.mkdtemp()

    file_ext   = os.path.splitext(uploaded.name)[1]
    excel_path = os.path.join(st.session_state.tmpdir, f"input{file_ext}")

    if st.session_state.get("last_uploaded") != uploaded.name:
        with open(excel_path, "wb") as f:
            f.write(uploaded.read())
        st.session_state.last_uploaded = uploaded.name
        st.session_state.sheet_names   = get_sheet_names(excel_path)
        st.session_state.sheet_cache   = {}
        st.session_state.selected_idx  = 0
        st.session_state.focus_field   = None
        # Clear any cached renders from previous upload
        for key in list(st.session_state.keys()):
            if key.startswith("_rendered_"):
                del st.session_state[key]

    with col_sheet_dropdown:
        st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)
        selected_sheet = st.selectbox(
            "Sheet", st.session_state.sheet_names, label_visibility="collapsed"
        )

    st.markdown("<hr style='border-color:#30363d; margin-top:-10px;'>", unsafe_allow_html=True)

    if selected_sheet not in st.session_state.sheet_cache:
        with st.spinner(f"Reading '{selected_sheet}' directly from Excel..."):
            data, sheet_type, report_header = extract_from_excel(excel_path, selected_sheet)
            st.info(f"Detected Sheet Type: {sheet_type}")
            if not data:
                st.warning(f"No data found in sheet '{selected_sheet}'.")
                st.stop()
            st.session_state.sheet_cache[selected_sheet] = {
    "data": data,
    "header": report_header
}
            st.session_state.selected_idx = 0
            st.session_state.focus_field  = None

    active = st.session_state.sheet_cache[selected_sheet]
    data   = active["data"]

    if st.session_state.selected_idx >= len(data):
        st.session_state.selected_idx = 0

    curr_claim        = data[st.session_state.selected_idx]
    col_nav, col_main = st.columns([1.2, 3.8], gap="large")

    # ── LEFT PANEL ────────────────────────────────────────────────────────
    with col_nav:
        with st.container(height=600, border=False):
            st.markdown("<p style='color:#8b949e; font-weight:bold; font-size:12px; text-transform:uppercase;'>TPA Records</p>", unsafe_allow_html=True)
            for i, row_data in enumerate(data):
                is_sel   = "selected-card" if st.session_state.selected_idx == i else ""
                c_id     = detect_claim_id(row_data, i)
                c_name   = get_val(row_data, ["Insured Name", "Name", "Company", "Claimant", "TPA_NAME"], "Unknown Entity")
                raw_st   = get_val(row_data, ["Status", "CLAIM_STATUS"], "")
                c_status = raw_st or ("Yet to Review" if i == 0 else "In Progress" if i == 1 else "Submitted")
                s_cls    = "status-progress" if "progress" in c_status.lower() or c_status.lower() == "open" else "status-text"
                st.markdown(f"""
                <div class="claim-card {is_sel}">
                    <div style="font-weight:bold;color:white;font-size:15px;">{c_id}</div>
                    <div style="color:#8b949e;font-size:13px;margin-top:2px;">{c_name}</div>
                    <div class="{s_cls}">{c_status}</div>
                </div>""", unsafe_allow_html=True)
                if st.button("Select", key=f"sel_{selected_sheet}_{i}", use_container_width=True):
                    st.session_state.selected_idx = i
                    st.session_state.focus_field  = None
                    st.rerun()

    # ── RIGHT PANEL ───────────────────────────────────────────────────────
    with col_main:
        head_left, head_right = st.columns([3, 1])
        curr_claim_id = detect_claim_id(curr_claim)

        with head_left:
            st.markdown("<p style='color:#8b949e;font-weight:bold;font-size:12px;text-transform:uppercase;'>Review Details</p>", unsafe_allow_html=True)
            sheet_header = active.get("header", "")
            h_date   = get_val(curr_claim, ["Loss Date", "Date", "LOSS_DATE"],                                 "N/A")
            h_status = get_val(curr_claim, ["Status", "CLAIM_STATUS"],                                         "Submitted")
            h_total  = get_val(curr_claim, ["Total Incurred", "Incurred", "Total", "Amount", "TOTAL_INCURRED"], "$0")
            st.markdown(f"""
                <div class="mid-header-title">{sheet_header}</div>
                
            """, unsafe_allow_html=True)

        with head_right:
            st.markdown("<p style='color:#8b949e;font-weight:bold;font-size:12px;text-transform:uppercase;text-align:right;'>Export Selection</p>", unsafe_allow_html=True)
            b1, b2 = st.columns(2)
            with b1:
                if st.button("☑ All", key=f"all_{selected_sheet}_{curr_claim_id}", use_container_width=True):
                    for f in curr_claim:
                        st.session_state[f"chk_{selected_sheet}_{curr_claim_id}_{f}"] = True
                    st.rerun()
            with b2:
                if st.button("☐ None", key=f"none_{selected_sheet}_{curr_claim_id}", use_container_width=True):
                    for f in curr_claim:
                        st.session_state[f"chk_{selected_sheet}_{curr_claim_id}_{f}"] = False
                    st.rerun()

        st.markdown("<hr style='border-color:#30363d;margin-top:8px;'>", unsafe_allow_html=True)

        

        hc = st.columns([2, 2.6, 2.6, 0.8, 0.5])
        with hc[0]: st.markdown("**FIELD**")
        with hc[1]: st.markdown("**EXTRACTED VALUE**")
        with hc[2]: st.markdown("**MODIFIED VALUE**")

        row_index = st.session_state.selected_idx

        for field, info in curr_claim.items():

            ek = f"edit_{selected_sheet}_{row_index}_{field}"
            xk = f"chk_{selected_sheet}_{row_index}_{field}"
            mk = f"mod_{selected_sheet}_{row_index}_{field}"

            if ek not in st.session_state: st.session_state[ek] = False
            if xk not in st.session_state: st.session_state[xk] = True
            if mk not in st.session_state: st.session_state[mk] = info.get("modified", info["value"])

            cl, co, cm, cx = st.columns([2, 2.6, 2.6, 0.5], gap="small")

            with cl:
                st.markdown(
                    f"<div style='height:40px;display:flex;align-items:center;"
                    f"color:#c9d1d9;font-size:12px;font-weight:bold;text-transform:uppercase;'>"
                    f"{field}</div>", unsafe_allow_html=True)

            with co:
                st.text_input("o", value=info["value"],
                              key=f"orig_{selected_sheet}_{row_index}_{field}",
                              label_visibility="collapsed", disabled=True)

            with cm:
                nv = st.text_input("m", key=mk, label_visibility="collapsed")
                st.session_state.sheet_cache[selected_sheet]["data"][
                    st.session_state.selected_idx][field]["modified"] = nv

        




            with cx:
                st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
                st.checkbox("", key=xk, label_visibility="collapsed")

        st.markdown("<hr style='border-color:#30363d;margin-top:12px;'>", unsafe_allow_html=True)

        _, btn_col = st.columns([4, 1.5])
        with btn_col:
            # if st.button(f"☑ Export Sheet '{selected_sheet}' to JSON",
            #              type="primary", use_container_width=True):
            #     export_data = {}
            #     for i, row in enumerate(data):
            #         c_id = detect_claim_id(row, i)
            #         rec  = {}
            #         for fld, inf in row.items():
            #             if st.session_state.get(f"chk_{selected_sheet}_{c_id}_{fld}", True):
            #                 mod  = inf.get("modified", "")
            #                 orig = inf.get("value", "")
            #                 rec[fld] = mod if mod != orig else orig
            #         export_data[c_id] = clean_duplicate_fields(rec)

            #     saved    = save_feature_store(selected_sheet, export_data)
            #     json_str = json.dumps(export_data, indent=2)
            #     st.success(f"✅ Saved → {saved}")
            #     st.download_button("📥 Download JSON", data=json_str,
            #                        file_name=f"{selected_sheet}_validated.json",
            #                        mime="application/json", use_container_width=True)
            selected_export_sheets = st.multiselect(
    "Select Sheets To Export",
    st.session_state.sheet_names,
    default=[selected_sheet]
)
            if st.button("☑ Export Selected Sheets to JSON", type="primary"):

                export_data = {}

                for sheet in selected_export_sheets:

                    sheet_data = st.session_state.sheet_cache[sheet]["data"]
                    sheet_header = st.session_state.sheet_cache[sheet].get("header", "")

                    records = []

                    for i, row in enumerate(sheet_data):

                        record = {}

                        for fld, inf in row.items():

                            val = inf.get("modified") if inf.get("modified") != inf.get("value") else inf.get("value")

                            record[fld] = {
                    "value": val,
                    "row": inf.get("excel_row"),
                    "column": inf.get("excel_col")
                }

                        records.append(record)

                    export_data[sheet] = {
            "report_header": sheet_header,
            "records": records
        }

                json_str = json.dumps(export_data, indent=2)

                st.download_button(
        "📥 Download JSON",
        data=json_str,
        file_name="loss_run_export.json",
        mime="application/json"
    )
    #         if st.button("☑ Export Selected Sheets to JSON", type="primary"):

    #             export_data = {}

    #             for sheet in selected_export_sheets:

    #                 sheet_data = st.session_state.sheet_cache[sheet]["data"]
    #                 sheet_export = []

    #                 for i, row in enumerate(sheet_data):

    #                     record = {}

    #                     for fld, inf in row.items():

    #                         val = inf.get("modified") if inf.get("modified") != inf.get("value") else inf.get("value")

    #                         record[fld] = {
    #         "value": val,
    #         "row": inf.get("excel_row"),
    #         "column": inf.get("excel_col")
    #     }

    #                         sheet_export.append(record)

    # #                 sheet_export = {}

    # #                 for i, row in enumerate(sheet_data):

    # #                     claim_id = detect_claim_id(row, i)

    # #                     record = {}

    # #                     for fld, inf in row.items():

    # #                         val = inf.get("modified") if inf.get("modified") != inf.get("value") else inf.get("value")

    # #                         record[fld] = {
    # #     "value": val,
    # #     "row": inf.get("excel_row"),
    # #     "column": inf.get("excel_col")
    # # }
                    


    #                 export_data[sheet] = sheet_export

    #             json_str = json.dumps(export_data, indent=2)

    #             st.download_button(
    #     "📥 Download JSON",
    #     data=json_str,
    #     file_name="loss_run_export.json",
    #     mime="application/json"
    # )