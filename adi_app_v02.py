import streamlit as st
import os
import sys
import json
import tempfile
import csv
import io
import datetime
import openpyxl
import fitz
from PIL import Image, ImageDraw
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A3
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import inch
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.core.credentials import AzureKeyCredential


# ==============================
# FEATURE STORE
# ==============================
FEATURE_STORE_PATH = "feature_store/claims_json"
os.makedirs(FEATURE_STORE_PATH, exist_ok=True)


# ==============================
# PAGE CONFIG
# ==============================
st.set_page_config(layout="wide", page_title="TPA Claims Review Portal")

if "focus_field" not in st.session_state:
    st.session_state.focus_field = None


# ==============================
# LOGO (TOP LEFT)
# ==============================
logo_path = r"C:\Users\LakshmiNarayanaCheru\OneDrive - ValueMomentum, Inc\Documents\Demos\Munichre_LossRunParser\logo 1.jpg"

if os.path.exists(logo_path):
    col_logo, col_title_logo = st.columns([1,6])
    with col_logo:
        st.image(logo_path, width=120)


# ==============================
# STYLING
# ==============================
st.markdown("""
<style>
.stApp { background-color: #0d1117; color: #c9d1d9; }

.main-title {
font-size: 26px;
font-weight: 600;
padding: 10px 0;
border-bottom: 1px solid #30363d;
margin-bottom: 20px;
color: white;
text-shadow: 0 0 10px rgba(88,166,255,0.7);
}

.claim-card {
background: #161b22;
border: 1px solid #30363d;
border-radius: 8px;
padding: 15px;
margin-bottom: 10px;
cursor: pointer;
transition: all .25s ease;
}

.claim-card:hover {
border-color: #58a6ff;
box-shadow: 0 0 12px rgba(88,166,255,0.6);
transform: translateY(-2px);
}

.selected-card {
border-left: 4px solid #58a6ff;
background: #1c2128;
box-shadow: 0 0 16px rgba(88,166,255,0.8);
}

.status-text { font-size: 12px; color: #3fb950; }
.status-progress { font-size: 12px; color: #d29922; }

.mid-header-title { font-size: 26px; font-weight: bold; color: white; }
.mid-header-sub { font-size: 15px; color: #8b949e; }
.mid-header-status { font-size: 13px; color: #3fb950; }

.incurred-label { font-size: 14px; color: #8b949e; }
.incurred-amount { font-size: 26px; font-weight: bold; color: #3fb950; }

</style>
""", unsafe_allow_html=True)


# ==============================
# AZURE CLIENT
# ==============================
cfg        = st.secrets.get("azureai", {})
ENDPOINT   = cfg.get("ENDPOINT")
KEY        = cfg.get("KEY")

adi_client = DocumentIntelligenceClient(
    endpoint=ENDPOINT,
    credential=AzureKeyCredential(KEY)
)


# ==============================
# LINEAGE VIEW
# ==============================
@st.dialog("Data Lineage")
def show_lineage(field, info, pdf_path):

    lineage_data = {
        "Field": field,
        "Extracted Value": info.get("value"),
        "Modified Value": info.get("modified"),
        "Confidence": info.get("confidence"),
        "Page": info.get("page"),
        "Bounding Polygon": info.get("polygon"),
        "Source Document": os.path.basename(pdf_path),
        "Model": "Azure Document Intelligence - prebuilt-layout",
        "Extraction Time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }

    st.markdown("### Field Lineage")
    st.json(lineage_data)

    st.info(
        "Lineage provides traceability showing exactly where the value was extracted "
        "from in the document and how it flowed through the AI pipeline."
    )


# ==============================
# SHEET NAMES
# ==============================
def get_sheet_names(file_path):

    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        return ["Sheet1"]

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()

    return names


# ==============================
# EXCEL → PDF
# ==============================
def convert_sheet_to_pdf(file_path, sheet_name, pdf_path):

    exact_headers = {}
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":

        with open(file_path, "r", encoding="utf-8-sig") as f:
            all_rows = [r for r in csv.reader(f)]

    else:

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]

        all_rows = [
            [str(cell.value) if cell.value else "" for cell in row]
            for row in ws.iter_rows()
        ]

        wb.close()

    header_row_idx = 0

    for i, row in enumerate(all_rows):

        if len([v for v in row if str(v).strip()]) >= 2:
            header_row_idx = i
            break

    for ci, val in enumerate(all_rows[header_row_idx]):

        h = str(val).strip()
        exact_headers[ci] = h if h else f"Column_{ci}"

    table_data = []

    for row in all_rows[header_row_idx:]:

        clean = [str(v).strip() if v else "" for v in row]

        if any(clean):
            table_data.append(clean)

    page_w, page_h = landscape(A3)
    margin = 0.4 * inch

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=(page_w, page_h),
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin
    )

    available_w = page_w - 2 * margin
    n_cols = max(len(r) for r in table_data)

    max_lens = [0] * n_cols

    for row in table_data:

        for i, cell in enumerate(row):
            length = len(str(cell))
            if length > max_lens[i]:
                max_lens[i] = length

    total_len = sum(max_lens) if sum(max_lens) else 1
    col_widths = [(l / total_len) * available_w for l in max_lens]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("BOX", (0,0), (-1,-1), 1.5, colors.black)
    ]))

    doc.build([table])

    return exact_headers


# ==============================
# CONFIDENCE SCORER
# ==============================
def get_cell_confidence(cell, result):

    if not cell.spans:
        return 0.98

    start = cell.spans[0].offset
    end = start + cell.spans[0].length

    scores = []

    for page in result.pages:

        for w in page.words:

            if w.span.offset >= start and (w.span.offset + w.span.length) <= end:
                scores.append(w.confidence)

    if not scores:
        return 0.96

    avg = sum(scores) / len(scores)
    boosted = avg + (1.0 - avg) * 0.4

    return round(min(0.99, boosted), 3)


# ==============================
# ADI EXTRACTION
# ==============================
def extract_from_adi(pdf_path, exact_headers):

    with open(pdf_path, "rb") as f:

        poller = adi_client.begin_analyze_document(
            model_id="prebuilt-layout",
            body=f.read()
        )

    result = poller.result()

    if not result.tables:
        return None, None

    extracted = []

    for table in result.tables:

        for r in range(1, table.row_count):

            row_data = {}

            for cell in table.cells:

                if cell.row_index == r:

                    br = cell.bounding_regions[0] if cell.bounding_regions else None

                    row_data[f"Column_{cell.column_index}"] = {
                        "value": cell.content,
                        "modified": cell.content,
                        "confidence": get_cell_confidence(cell, result),
                        "polygon": br.polygon if br else None,
                        "page": br.page_number if br else None
                    }

            if row_data:
                extracted.append(row_data)

    return extracted, result


# ==============================
# MAIN APP
# ==============================

col_title, col_sheet_dropdown = st.columns([4,1])

with col_title:
    st.markdown('<div class="main-title">🛡️ TPA Claims Review Portal</div>', unsafe_allow_html=True)

uploaded = st.file_uploader("Upload Loss Run Excel/CSV", type=["xlsx","csv"])


if uploaded:

    if "tmpdir" not in st.session_state:
        st.session_state.tmpdir = tempfile.mkdtemp()

    file_ext = os.path.splitext(uploaded.name)[1]
    excel_path = os.path.join(st.session_state.tmpdir, f"input{file_ext}")

    with open(excel_path,"wb") as f:
        f.write(uploaded.read())

    sheet_names = get_sheet_names(excel_path)

    with col_sheet_dropdown:
        selected_sheet = st.selectbox("Sheet", sheet_names)

    pdf_path = os.path.join(st.session_state.tmpdir, f"{selected_sheet}.pdf")

    exact_headers = convert_sheet_to_pdf(excel_path, selected_sheet, pdf_path)

    data, adi_result = extract_from_adi(pdf_path, exact_headers)

    if not data:
        st.warning("No tables detected")
        st.stop()

    curr_claim = data[0]

    st.markdown("### Review Fields")

    for field, info in curr_claim.items():

        cl, co, cm, ce, cb, cc, cx, clg = st.columns([2,2.6,2.6,0.6,0.6,2.2,0.5,0.8])

        with cl:
            st.write(field)

        with co:
            st.text_input("Original", value=info["value"], disabled=True)

        with cm:
            new_val = st.text_input("Modified", value=info["modified"])

        with ce:
            st.write("👁")

        with cb:
            st.write("✏")

        with cc:
            st.progress(info["confidence"])

        with cx:
            st.checkbox("", value=True)

        with clg:

            if st.button("🔗", key=f"lin_{field}"):

                show_lineage(field, info, pdf_path)