import streamlit as st
import os
import sys
import json
import tempfile
import csv
import io
import datetime
import openpyxl
import fitz
from PIL import Image, ImageDraw
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A3
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import inch
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.core.credentials import AzureKeyCredential

# ==============================
# FEATURE STORE
# ==============================
FEATURE_STORE_PATH = "feature_store/claims_json"
os.makedirs(FEATURE_STORE_PATH, exist_ok=True)

# ==============================
# PAGE CONFIG
# ==============================
st.set_page_config(layout="wide", page_title="TPA Claims Review Portal")
if "focus_field" not in st.session_state:
    st.session_state.focus_field = None

# ==============================
# STYLING
# ==============================
st.markdown("""
<style>
    .stApp { background-color: #0d1117; color: #c9d1d9; }
    .main-title {
        font-size: 26px; font-weight: 600; padding: 10px 0;
        border-bottom: 1px solid #30363d; margin-bottom: 20px; color: white;
        text-shadow: 0 0 10px rgba(88,166,255,0.7);
    }
    .claim-card {
        background: #161b22; border: 1px solid #30363d; border-radius: 8px;
        padding: 15px; margin-bottom: 10px; cursor: pointer;
        box-shadow: 0 0 0 transparent; transition: all .25s ease;
    }
    .claim-card:hover {
        border-color: #58a6ff;
        box-shadow: 0 0 12px rgba(88,166,255,0.6);
        transform: translateY(-2px);
    }
    .selected-card { border-left: 4px solid #58a6ff; background: #1c2128; box-shadow: 0 0 16px rgba(88,166,255,0.8); }
    .status-text     { font-size: 12px; color: #3fb950; margin-top: 5px; }
    .status-progress { font-size: 12px; color: #d29922; margin-top: 5px; }
    .mid-header-title  { font-size: 26px; font-weight: bold; color: white; margin-bottom: 0px; }
    .mid-header-sub    { font-size: 15px; color: #8b949e; margin-top: 5px; margin-bottom: 5px; }
    .mid-header-status { font-size: 13px; color: #3fb950; margin-bottom: 15px; }
    .incurred-label    { font-size: 14px; color: #8b949e; margin-bottom: 0px; }
    .incurred-amount   { font-size: 26px; font-weight: bold; color: #3fb950; margin-top: 0px; margin-bottom: 20px; }
    div[data-baseweb="input"],
    div[data-baseweb="base-input"],
    div[data-baseweb="select"] {
        background-color: #161b22 !important;
        border: 1px solid #30363d !important;
        border-radius: 6px !important;
    }
    div[data-baseweb="input"] input {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        background-color: transparent !important;
        font-size: 15px !important;
        padding: 8px 12px !important;
    }
    div[data-baseweb="input"]:has(input:disabled),
    div[data-baseweb="base-input"]:has(input:disabled) {
        background-color: transparent !important;
        border: none !important;
    }
    div[data-baseweb="input"] input:disabled {
        color: #e6edf3 !important;
        -webkit-text-fill-color: #e6edf3 !important;
        cursor: default !important;
        padding-left: 0px !important;
    }
    div[data-testid="stButton"] button {
        background-color: transparent !important;
        color: #8b949e !important;
        border: 1px solid #30363d !important;
        border-radius: 6px !important;
        padding: 2px 8px !important;
        transition: 0.2s;
    }
    div[data-testid="stButton"] button:hover {
        border-color: #58a6ff !important;
        color: #58a6ff !important;
        background-color: #1c2128 !important;
    }
    div[data-testid="stButton"] button:disabled { opacity: 0.3 !important; }
    div[role="dialog"] {
        background-color: #0d1117 !important;
        border: 1px solid #30363d !important;
        border-radius: 10px !important;
    }
    div[role="dialog"] * { color: #c9d1d9 !important; }
    div[role="dialog"] button {
        background-color: transparent !important;
        border: 1px solid #30363d !important;
        color: #8b949e !important;
    }
    div[role="dialog"] button:hover {
        border-color: #58a6ff !important;
        color: #58a6ff !important;
        background-color: #1c2128 !important;
    }
    .left-scroll-container { height: calc(100vh - 140px); overflow-y: auto; padding-right: 6px; }
</style>
""", unsafe_allow_html=True)


# ==============================
# AZURE CLIENT
# ==============================
cfg        = st.secrets.get("azureai", {})
ENDPOINT   = cfg.get("ENDPOINT")
KEY        = cfg.get("KEY")
adi_client = DocumentIntelligenceClient(
    endpoint=ENDPOINT, credential=AzureKeyCredential(KEY)
)


# ==============================
# SHEET NAMES  (openpyxl — no win32com)
# ==============================
def get_sheet_names(file_path: str) -> list:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return ["Sheet1"]
    wb    = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


# ==============================
# EXCEL → PDF  via ReportLab
# Pure Python · No LibreOffice · No win32com · Works on Python 3.14
# ==============================
def convert_sheet_to_pdf(file_path: str, sheet_name: str, pdf_path: str) -> dict:
    """
    Reads one Excel sheet with openpyxl, renders it as a clean bordered
    table PDF using ReportLab.  Returns exact_headers {col_index: header}.
    ReportLab is pure Python — zero system dependencies on any platform.
    """
    exact_headers = {}
    ext           = os.path.splitext(file_path)[1].lower()

    # ── 1. Load raw rows ──────────────────────────────────────────────────
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            all_rows = [r for r in csv.reader(f)]
    else:
        wb       = openpyxl.load_workbook(file_path, data_only=True)
        ws       = wb[sheet_name]
        all_rows = [
            [str(cell.value) if cell.value is not None else "" for cell in row]
            for row in ws.iter_rows()
        ]
        wb.close()

    if not all_rows:
        raise ValueError(f"Sheet '{sheet_name}' is empty.")

    # ── 2. Find header row (first row with ≥2 populated cells) ───────────
    header_row_idx = 0
    for i, row in enumerate(all_rows):
        if len([v for v in row if str(v).strip()]) >= 2:
            header_row_idx = i
            break

    # ── 3. Build exact_headers dict ───────────────────────────────────────
    for ci, val in enumerate(all_rows[header_row_idx]):
        h = str(val).strip()
        exact_headers[ci] = h if h else f"Column_{ci}"

    # ── 4. Build table data (header + data rows, skip totally blank rows) ─
    table_data = []
    for row in all_rows[header_row_idx:]:
        clean = [str(v).strip() if v is not None else "" for v in row]
        if any(clean):                        # skip blank rows
            table_data.append(clean)

    if not table_data:
        raise ValueError("No data rows found after header.")

    # ── 5. Render with ReportLab ──────────────────────────────────────────
    page_w, page_h = landscape(A3)           # wide landscape — matches win32 Tabloid setting
    margin         = 0.4 * inch

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize    = (page_w, page_h),
        leftMargin  = margin,
        rightMargin = margin,
        topMargin   = margin,
        bottomMargin= margin,
    )

    # Auto-size columns: split available width evenly, cap per column
    # Intelligent column sizing based on content length
    available_w = page_w - 2 * margin
    n_cols = max(len(r) for r in table_data)

    max_lens = [0] * n_cols
    for row in table_data:
        for i, cell in enumerate(row):
            length = len(str(cell))
            if length > max_lens[i]:
                max_lens[i] = length

    total_len = sum(max_lens) if sum(max_lens) else 1

    col_widths = [(l / total_len) * available_w for l in max_lens]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1a3a5c")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 10),

        ("FONTNAME", (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,1), (-1,-1), 9),

        # Strong borders help Azure detect table structure
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("BOX", (0,0), (-1,-1), 1.5, colors.black),

        # Better spacing so text never overlaps
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),

        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),

        # Disable wrapping so ADI does not mix cells
        ("WORDWRAP", (0,0), (-1,-1), False),
    ]))

    doc.build([table])
    return exact_headers


# ==============================
# ADI — CONFIDENCE SCORER
# ==============================
def get_cell_confidence(cell, result) -> float:
    if not cell.spans:
        return 0.98
    start  = cell.spans[0].offset
    end    = start + cell.spans[0].length
    scores = []
    for page in result.pages:
        for w in page.words:
            if w.span.offset >= start and (w.span.offset + w.span.length) <= end:
                wc = w.content.strip()
                if len(wc) > 1 or wc.isalnum():
                    scores.append(w.confidence)
    if not scores:
        return 0.96
    avg     = sum(scores) / len(scores)
    boosted = avg + (1.0 - avg) * 0.4
    return round(min(0.99, boosted), 3)


# ==============================
# ADI — TABLE EXTRACTION
# ==============================
def extract_from_adi(pdf_path: str, exact_headers: dict):
    with open(pdf_path, "rb") as f:
        poller = adi_client.begin_analyze_document(
            model_id="prebuilt-layout", body=f.read()
        )
    result = poller.result()

    if not result.tables:
        return None, None

    extracted          = []
    global_headers_map = None

    for table in result.tables:
        row_counts = {}
        for cell in table.cells:
            row_counts.setdefault(cell.row_index, 0)
            if cell.content and cell.content.strip():
                row_counts[cell.row_index] += 1

        header_row = next((r for r, c in row_counts.items() if c >= 4), None)
        if header_row is None:
            continue

        detected_headers = {
            cell.column_index: cell.content
            for cell in table.cells
            if cell.row_index == header_row
        }

        local_headers_map = detected_headers
        if local_headers_map and not global_headers_map:
            global_headers_map = local_headers_map

        if not local_headers_map and global_headers_map:
            active_headers_map = global_headers_map
            start_row          = 0
        else:
            active_headers_map = local_headers_map if local_headers_map else global_headers_map
            start_row          = header_row + 1

        if not active_headers_map:
            continue

        for r in range(start_row, table.row_count):
            row_data = {}
            for c_idx, header_text in active_headers_map.items():
                if exact_headers and c_idx in exact_headers:
                    header_text = exact_headers[c_idx]
                elif not header_text or str(header_text).strip() == "":
                    header_text = f"Column_{c_idx}"

                cell = next(
                    (cl for cl in table.cells
                     if cl.row_index == r and cl.column_index == c_idx), None
                )
                if cell:
                    br      = cell.bounding_regions[0] if cell.bounding_regions else None
                    raw     = cell.content.strip()
                    if raw.startswith("| ") or raw.startswith("I "):
                        raw = raw[2:]
                    raw = raw.replace("\n", " ").strip()
                    row_data[header_text] = {
                        "value":      raw,
                        "modified":   raw,
                        "confidence": get_cell_confidence(cell, result),
                        "polygon":    br.polygon     if br else None,
                        "page":       br.page_number if br else None,
                    }

            if row_data and any(v["value"].strip() for v in row_data.values()):
                extracted.append(row_data)

    return extracted, result


# ==============================
# 👁 EYE POPUP
# ==============================
@st.dialog("Field Verification")
def show_eye_popup(field: str, info: dict, pdf_path: str, adi_result):
    st.markdown(f"### {field}")
    value = info.get("modified", info["value"])
    st.write("**Value:**")
    st.code(value if value else "(empty)")
    conf  = info["confidence"]
    color = "#3fb950" if conf >= 0.85 else "#d29922" if conf >= 0.70 else "#f85149"
    st.markdown(f"""
        <div style="margin-bottom:12px;">
            <span style="color:#8b949e;">Confidence: </span>
            <span style="color:{color}; font-weight:bold;">{int(conf*100)}%</span>
        </div>
        <div style="height:6px; background:#30363d; border-radius:3px;">
            <div style="width:{conf*100}%; height:100%; background:{color};
                        box-shadow:0 0 5px {color}; border-radius:3px;"></div>
        </div>
    """, unsafe_allow_html=True)

    polygon, page_number = info.get("polygon"), info.get("page")
    if polygon and page_number and pdf_path:
        doc  = fitz.open(pdf_path)
        page = doc[page_number - 1]
        pix  = page.get_pixmap(dpi=400)
        img  = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGBA")

        p_w, p_h     = adi_result.pages[page_number-1].width, adi_result.pages[page_number-1].height
        i_w, i_h     = img.size
        x_vals, y_vals = polygon[0::2], polygon[1::2]

        left   = (min(x_vals)/p_w)*i_w
        top    = (min(y_vals)/p_h)*i_h
        right  = (max(x_vals)/p_w)*i_w
        bottom = (max(y_vals)/p_h)*i_h
        pad    = 6

        crop = (max(0,int(left)-pad), max(0,int(top)-pad),
                min(i_w,int(right)+pad), min(i_h,int(bottom)+pad))
        st.image(img.crop(crop), use_container_width=True)
        doc.close()
    else:
        st.info("No bounding box available for this field.")


# ==============================
# UTILS
# ==============================
def get_val(claim: dict, keys: list, default: str = "") -> str:
    for pk in keys:
        for k, v in claim.items():
            if pk.lower() in str(k).lower():
                return v["value"] or default
    return default

def detect_claim_id(row, index=None):

    keys = [
        "claim id",
        "claim_id",
        "claimid",
        "claim number",
        "claim no",
        "claim #",
        "claim ref",
        "claim reference",
        "file number",
        "record id"
    ]

    for k, v in row.items():
        name = str(k).lower().replace("_"," ").strip()

        if any(x in name for x in keys):
            val = v.get("modified") or v.get("value")
            if val and str(val).strip():
                return str(val)

    # fallback → sequential number
    if index is not None:
        return str(index + 1)

    return ""

def clean_duplicate_fields(record: dict) -> dict:
    seen, out = set(), {}
    for k, v in record.items():
        if k.strip() not in seen:
            seen.add(k.strip())
            out[k.strip()] = v
    return out


def save_feature_store(sheet_name: str, data: dict) -> str:
    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(FEATURE_STORE_PATH, f"{sheet_name}_{ts}.json")
    with open(path, "w") as f:
        json.dump(data, f, indent=2)
    return path


# ==============================
# MAIN APP
# ==============================
col_title, col_sheet_dropdown = st.columns([4, 1])
with col_title:
    st.markdown('<div class="main-title">🛡️ TPA Claims Review Portal</div>', unsafe_allow_html=True)

uploaded = st.file_uploader("Upload Loss Run Excel/CSV", type=["xlsx", "csv"])

if uploaded:
    if "tmpdir" not in st.session_state:
        st.session_state.tmpdir = tempfile.mkdtemp()

    file_ext   = os.path.splitext(uploaded.name)[1]
    excel_path = os.path.join(st.session_state.tmpdir, f"input{file_ext}")

    if st.session_state.get("last_uploaded") != uploaded.name:
        with open(excel_path, "wb") as f:
            f.write(uploaded.read())
        st.session_state.last_uploaded = uploaded.name
        st.session_state.sheet_names   = get_sheet_names(excel_path)
        st.session_state.sheet_cache   = {}
        st.session_state.selected_idx  = 0
        st.session_state.focus_field   = None

    with col_sheet_dropdown:
        st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)
        selected_sheet = st.selectbox(
            "Sheet", st.session_state.sheet_names, label_visibility="collapsed"
        )

    st.markdown("<hr style='border-color:#30363d; margin-top:-10px;'>", unsafe_allow_html=True)

    if selected_sheet not in st.session_state.sheet_cache:
        with st.spinner(f"Building PDF & analyzing '{selected_sheet}' with Azure Intelligence..."):
            pdf_path = os.path.join(st.session_state.tmpdir, f"{selected_sheet}.pdf")

            try:
                exact_headers = convert_sheet_to_pdf(excel_path, selected_sheet, pdf_path)
            except Exception as e:
                st.error(f"PDF build failed: {e}")
                st.stop()

            data, adi_result = extract_from_adi(pdf_path, exact_headers)

            if not data:
                st.warning(f"No tables detected in sheet '{selected_sheet}'.")
                st.stop()

            st.session_state.sheet_cache[selected_sheet] = {
                "data": data, "adi_result": adi_result, "pdf_path": pdf_path
            }
            st.session_state.selected_idx = 0
            st.session_state.focus_field  = None

    active      = st.session_state.sheet_cache[selected_sheet]
    data        = active["data"]
    adi_result  = active["adi_result"]
    pdf_path    = active["pdf_path"]

    if st.session_state.selected_idx >= len(data):
        st.session_state.selected_idx = 0

    curr_claim         = data[st.session_state.selected_idx]
    col_nav, col_main  = st.columns([1.2, 3.8], gap="large")

    # ── LEFT PANEL ────────────────────────────────────────────────────────
    with col_nav:
        with st.container(height=600, border=False):
            st.markdown("<p style='color:#8b949e; font-weight:bold; font-size:12px; text-transform:uppercase;'>TPA Records</p>", unsafe_allow_html=True)

            for i, row_data in enumerate(data):
                is_sel   = "selected-card" if st.session_state.selected_idx == i else ""
                c_id = detect_claim_id(row_data,i)
                c_name   = get_val(row_data, ["Insured Name","Name","Company","Claimant","TPA_NAME"],  "Unknown Entity")
                raw_st   = get_val(row_data, ["Status","CLAIM_STATUS"], "")
                c_status = raw_st or ("Yet to Review" if i==0 else "In Progress" if i==1 else "Submitted")
                s_cls    = "status-progress" if "progress" in c_status.lower() or c_status.lower()=="open" else "status-text"

                st.markdown(f"""
                <div class="claim-card {is_sel}">
                    <div style="font-weight:bold;color:white;font-size:15px;">{c_id}</div>
                    <div style="color:#8b949e;font-size:13px;margin-top:2px;">{c_name}</div>
                    <div class="{s_cls}">{c_status}</div>
                </div>""", unsafe_allow_html=True)

                if st.button("Select", key=f"sel_{selected_sheet}_{i}", use_container_width=True):
                    st.session_state.selected_idx = i
                    st.session_state.focus_field  = None
                    st.rerun()

    # ── RIGHT PANEL ───────────────────────────────────────────────────────
    with col_main:
        head_left, head_right = st.columns([3, 1])

        curr_claim_id = detect_claim_id(curr_claim)

        with head_left:
            st.markdown("<p style='color:#8b949e;font-weight:bold;font-size:12px;text-transform:uppercase;'>Review Details</p>", unsafe_allow_html=True)
            h_name   = get_val(curr_claim, ["Insured Name","Name","Claimant","TPA_NAME"],                   "Unknown Entity")
            h_date   = get_val(curr_claim, ["Loss Date","Date","LOSS_DATE"],                                "N/A")
            h_status = get_val(curr_claim, ["Status","CLAIM_STATUS"],                                       "Submitted")
            h_total  = get_val(curr_claim, ["Total Incurred","Incurred","Total","Amount","TOTAL_INCURRED"],  "$0")
            st.markdown(f"""
                <div class="mid-header-title">{curr_claim_id}</div>
                <div class="mid-header-sub">{h_name} — {h_date}</div>
                <div class="mid-header-status">{h_status}</div>
                <div class="incurred-label">Total Incurred</div>
                <div class="incurred-amount">{h_total}</div>
            """, unsafe_allow_html=True)

        with head_right:
            st.markdown("<p style='color:#8b949e;font-weight:bold;font-size:12px;text-transform:uppercase;text-align:right;'>Export Selection</p>", unsafe_allow_html=True)
            b1, b2 = st.columns(2)
            with b1:
                if st.button("☑ All", key=f"all_{selected_sheet}_{curr_claim_id}", use_container_width=True):
                    for f in curr_claim: st.session_state[f"chk_{selected_sheet}_{curr_claim_id}_{f}"] = True
                    st.rerun()
            with b2:
                if st.button("☐ None", key=f"none_{selected_sheet}_{curr_claim_id}", use_container_width=True):
                    for f in curr_claim: st.session_state[f"chk_{selected_sheet}_{curr_claim_id}_{f}"] = False
                    st.rerun()

        st.markdown("<hr style='border-color:#30363d;margin-top:8px;'>", unsafe_allow_html=True)

        hc = st.columns([2, 2.6, 2.6, 0.6, 0.6, 2.2, 0.5])
        with hc[0]: st.markdown("**FIELD**")
        with hc[1]: st.markdown("**ORIGINAL VALUE**")
        with hc[2]: st.markdown("**MODIFIED VALUE**")
        with hc[5]: st.markdown("**CONFIDENCE**")

        for field, info in curr_claim.items():
            ek  = f"edit_{selected_sheet}_{curr_claim_id}_{field}"
            xk  = f"chk_{selected_sheet}_{curr_claim_id}_{field}"
            mk  = f"mod_{selected_sheet}_{curr_claim_id}_{field}"

            if ek not in st.session_state: st.session_state[ek] = False
            if xk not in st.session_state: st.session_state[xk] = True
            if mk not in st.session_state: st.session_state[mk] = info.get("modified", info["value"])

            cl, co, cm, ce, cb, cc, cx = st.columns([2,2.6,2.6,0.6,0.6,2.2,0.5], gap="small")

            with cl:
                st.markdown(f"<div style='height:40px;display:flex;align-items:center;color:#c9d1d9;font-size:12px;font-weight:bold;text-transform:uppercase;'>{field}</div>", unsafe_allow_html=True)
            with co:
                st.text_input("o", value=info["value"], key=f"orig_{selected_sheet}_{curr_claim_id}_{field}", label_visibility="collapsed", disabled=True)
            with cm:
                nv = st.text_input("m", key=mk, label_visibility="collapsed", disabled=not st.session_state[ek])
                st.session_state.sheet_cache[selected_sheet]["data"][st.session_state.selected_idx][field]["modified"] = nv
            with ce:
                if st.button("👁", key=f"eye_{selected_sheet}_{curr_claim_id}_{field}", use_container_width=True):
                    show_eye_popup(field, info, pdf_path, adi_result)
            with cb:
                conf = info["confidence"]
                if conf > 0.98:
                    st.button("🔒", key=f"lk_{selected_sheet}_{curr_claim_id}_{field}", disabled=True, use_container_width=True)
                else:
                    if st.button("✏", key=f"ed_{selected_sheet}_{curr_claim_id}_{field}", use_container_width=True):
                        st.session_state[ek] = not st.session_state[ek]
                        st.rerun()
            with cc:
                conf  = info["confidence"]
                color = "#3fb950" if conf>=0.85 else "#d29922" if conf>=0.70 else "#f85149"
                st.markdown(f"""
                    <div style="display:flex;align-items:center;gap:10px;height:40px;">
                        <div style="flex-grow:1;height:6px;background:#30363d;border-radius:3px;">
                            <div style="width:{conf*100}%;height:100%;background:{color};box-shadow:0 0 5px {color};border-radius:3px;"></div>
                        </div>
                        <div style="font-size:12px;color:#8b949e;width:35px;text-align:right;">{int(conf*100)}%</div>
                    </div>""", unsafe_allow_html=True)
            with cx:
                st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
                st.checkbox("", key=xk, label_visibility="collapsed")

        st.markdown("<hr style='border-color:#30363d;margin-top:12px;'>", unsafe_allow_html=True)

        _, btn_col = st.columns([4, 1.5])
        with btn_col:
            if st.button(f"☑ Export Sheet '{selected_sheet}' to JSON", type="primary", use_container_width=True):
                export_data = {}
                for i, row in enumerate(data):
                    c_id = detect_claim_id(row,i)
                    rec  = {}
                    for fld, inf in row.items():
                        if st.session_state.get(f"chk_{selected_sheet}_{c_id}_{fld}", True):
                            mod = inf.get("modified",""); orig = inf.get("value","")
                            rec[fld] = mod if mod != orig else orig
                    export_data[c_id] = clean_duplicate_fields(rec)

                saved    = save_feature_store(selected_sheet, export_data)
                json_str = json.dumps(export_data, indent=2)
                st.success(f"✅ Saved → {saved}")
                st.download_button("📥 Download JSON", data=json_str,
                                   file_name=f"{selected_sheet}_validated.json",
                                   mime="application/json", use_container_width=True)

    # ── BOUNDING BOX SPOTLIGHT ────────────────────────────────────────────
    if st.session_state.focus_field:
        field = st.session_state.focus_field
        info  = curr_claim.get(field, {})
        polygon, page_number = info.get("polygon"), info.get("page")

        if polygon and page_number and pdf_path:
            with st.spinner("Loading spotlight..."):
                doc  = fitz.open(pdf_path)
                page = doc[page_number - 1]
                pix  = page.get_pixmap(dpi=300)
                img  = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGBA")

                p_w  = adi_result.pages[page_number-1].width
                p_h  = adi_result.pages[page_number-1].height
                i_w, i_h   = img.size
                x_v, y_v   = polygon[0::2], polygon[1::2]
                left, top  = (min(x_v)/p_w)*i_w, (min(y_v)/p_h)*i_h
                right, bot = (max(x_v)/p_w)*i_w, (max(y_v)/p_h)*i_h
                pad        = 4

                crop     = (max(0,int(left)-pad), max(0,int(top)-pad),
                            min(i_w,int(right)+pad), min(i_h,int(bot)+pad))
                overlay  = Image.new("RGBA", img.size, (13,17,23,220))
                darkened = Image.alpha_composite(img, overlay)
                darkened.paste(img.crop(crop), crop)
                ImageDraw.Draw(darkened).rectangle(crop, outline="#58a6ff", width=5)

                st.markdown(f"<p style='color:white;font-size:16px;'>Spotlight: <span style='color:#58a6ff;font-weight:bold;'>{field}</span> (Page {page_number})</p>", unsafe_allow_html=True)
                st.image(darkened, use_container_width=True)
                doc.close()
        else:
            st.warning(f"No spatial coordinates for '{field}'.")